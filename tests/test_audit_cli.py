"""Tests for performance comparison command-line modules."""

# Python imports
from contextlib import redirect_stderr, redirect_stdout
import io
import json
import os
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile
import unittest
from unittest import mock

# Third-party imports
from openpyxl import load_workbook
import yaml

from ppar.errors import PpaError
from ppar.analytics import cli as _analytics_cli
from ppar.audit import compare_snapshots, write_audit_report_bundle
from ppar.audit.cli import site_report as _site_report
from ppar.audit.config_validation import validate_config
from ppar.audit.run_settings import (
    audit_settings as _audit_settings,
    resolve_settings as _resolve_audit_settings,
)

_RESTATEMENT_COMPARISON_PATH = Path(
    "tests/data/axys/validation/ppar_audit_restatement.yaml"
)
_PORTFOLIO_COMPARISON_PATH = Path(
    "ppar/setup_templates/axys_apx_audit/axys_apx_audit.yaml"
)
_PACKAGED_AXYS_APX_DATA_PATH = Path("ppar/setup_templates/axys_apx_audit")
_AXYS_SNAPSHOT_PATH = Path("tests/data/axys/snapshots")
_VALIDATE_BUNDLE_MODULE = "ppar.audit.cli.validate_bundle"
_VALIDATE_CONFIG_MODULE = "ppar.audit.cli.validate_config"
_VALIDATE_DEMO_MATRIX_SCRIPT = Path("scripts/validate_demo_matrix.py")
_SETUP_MODULE = "ppar.audit.cli.setup"
_SITE_REPORT_MODULE = "ppar.audit.cli.site_report"
_ANALYTICS_MODULE = "ppar.analytics.cli"
_PPAR_MODULE = "ppar.cli"
_DEMO_QUIET_PHRASES = (
    "Using quarterly reporting.",
    "Time:",
    "Analytics demo output written to:",
    "Report bundle written to:",
    "Bundle artifacts:",
    "Portfolio Audit Report",
    "Security Audit Report",
)


class TestAuditCli(unittest.TestCase):
    """Verify command-line report generation and validation commands."""

    def test_site_audit_returns_nonzero_for_output_row_limit(self) -> None:
        """The Audit CLI surfaces an oversized-report failure without success output."""
        stderr = io.StringIO()
        with mock.patch.object(
            _site_report,
            "run_report",
            side_effect=PpaError(
                "Audit output row limit exceeded. "
                "No files were written for the oversized report.",
                None,
            ),
        ):
            with redirect_stderr(stderr):
                exit_code = _site_report.main(["."], prog="ppar audit")

        self.assertEqual(exit_code, 1)
        self.assertIn("Report failed: Audit output row limit exceeded", stderr.getvalue())
        self.assertIn(
            "No files were written for the oversized report",
            stderr.getvalue(),
        )

    def test_audit_yaml_policy_and_operational_cli_options_resolve_consistently(
        self,
    ) -> None:
        """YAML controls policy while the CLI can change one run's presentation."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            _write_audit_run_settings(site_directory / "ppar.yaml")
            configuration = yaml.safe_load(
                (site_directory / "ppar.yaml").read_text(encoding="utf-8")
            )
            configuration["audit"].update(
                {
                    "output_directory": "configured_output",
                    "title": "Configured Audit",
                    "xlsx_output": False,
                    "exclude_suppressed": True,
                }
            )
            (site_directory / "ppar.yaml").write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )

            configured = _resolve_audit_settings(
                site_directory,
                _audit_settings(configuration, required=True),
                output_directory=None,
                title=None,
                exclude_suppressed=None,
                include_reconstruction_diagnostics=None,
                require_causal_attribution=None,
                include_workbook=None,
                include_html_output=None,
                expand_all_supporting_files=None,
            )
            arguments = _site_report._argument_parser(
                prog="python run_audit.py",
            ).parse_args(
                [
                    ".",
                    "--output-directory",
                    str(site_directory / "one_run"),
                    "--title",
                    "One Run",
                    "--xlsx-only",
                    "--expand-supporting-files",
                ]
            )
            include_workbook, include_html_output = _site_report._output_overrides(
                arguments
            )
            overridden = _resolve_audit_settings(
                site_directory,
                _audit_settings(configuration, required=True),
                output_directory=arguments.output_directory,
                title=arguments.title,
                exclude_suppressed=None,
                include_reconstruction_diagnostics=None,
                require_causal_attribution=None,
                include_workbook=include_workbook,
                include_html_output=include_html_output,
                expand_all_supporting_files=(
                    True if arguments.expand_supporting_files else None
                ),
            )

        self.assertEqual(
            configured.output_directory,
            site_directory / "configured_output",
        )
        self.assertEqual(configured.title, "Configured Audit")
        self.assertFalse(configured.include_workbook)
        self.assertTrue(configured.exclude_suppressed)
        self.assertEqual(overridden.output_directory, site_directory / "one_run")
        self.assertEqual(overridden.title, "One Run")
        self.assertTrue(overridden.include_workbook)
        self.assertFalse(overridden.include_html_output)
        self.assertTrue(overridden.exclude_suppressed)
        self.assertFalse(overridden.include_reconstruction_diagnostics)
        self.assertTrue(overridden.expand_all_supporting_files)

    def test_audit_run_settings_default_missing_and_reject_unknown_keys(self) -> None:
        """Omitted Audit settings default while unknown settings fail closed."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            config_path = site_directory / "ppar.yaml"
            _write_audit_run_settings(config_path)
            configuration = yaml.safe_load(config_path.read_text(encoding="utf-8"))
            configuration["audit"] = {}
            config_path.write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )
            settings = _resolve_audit_settings(
                site_directory,
                _audit_settings(configuration, required=True),
                output_directory=None,
                title=None,
                exclude_suppressed=None,
                include_reconstruction_diagnostics=None,
                require_causal_attribution=None,
                include_workbook=None,
                include_html_output=None,
                expand_all_supporting_files=None,
            )
            self.assertEqual(settings.output_directory, site_directory / "output")
            self.assertIsNone(settings.title)
            self.assertTrue(settings.include_workbook)
            self.assertTrue(settings.include_html_output)
            self.assertFalse(settings.exclude_suppressed)
            self.assertFalse(settings.include_reconstruction_diagnostics)
            self.assertFalse(settings.expand_all_supporting_files)
            self.assertFalse(settings.require_causal_attribution)

            configuration["audit"]["html_ouput"] = True
            config_path.write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(
                PpaError,
                "audit has unsupported keys: html_ouput",
            ):
                _audit_settings(configuration, required=True)

    def test_audit_cli_rejects_removed_policy_and_boolean_flags(self) -> None:
        """The public command does not retain the superseded override surface."""
        for retired_flag in (
            "--output",
            "--exclude_suppressed",
            "--include-reconstruction-diagnostics",
            "--xlsx-output",
            "--no-xlsx-output",
            "--html-output",
            "--no-html-output",
            "--exclude-suppressed",
            "--no-exclude-suppressed",
            "--reconstruction-diagnostics",
            "--no-reconstruction-diagnostics",
            "--expand-all-supporting-files",
            "--no-expand-all-supporting-files",
            "--require-causal-attribution",
            "--no-require-causal-attribution",
            "--allow-incomplete-yaml",
        ):
            with self.subTest(retired_flag=retired_flag):
                with redirect_stderr(io.StringIO()), self.assertRaises(SystemExit):
                    _site_report._argument_parser(
                        prog="python run_audit.py",
                    ).parse_args([".", retired_flag])

    def test_audit_cli_output_modes_are_mutually_exclusive(self) -> None:
        """A run selects at most one nonstandard output mode."""
        with redirect_stderr(io.StringIO()), self.assertRaises(SystemExit):
            _site_report._argument_parser(
                prog="python run_audit.py",
            ).parse_args([".", "--html-only", "--xlsx-only"])

    def test_analytics_cli_rejects_retired_flag_names(self) -> None:
        """Analytics accepts only CLI names corresponding to YAML settings."""
        for retired_flag in (
            "--output",
            "--minimum-acceptable-return",
            "--risk-free-rate",
        ):
            with self.subTest(retired_flag=retired_flag):
                with redirect_stderr(io.StringIO()), self.assertRaises(SystemExit):
                    _analytics_cli.script_run_settings(Path.cwd(), [retired_flag])

    def test_analytics_run_settings_use_defaults_then_cli_overrides(self) -> None:
        """Analytics resolves omitted settings from documented defaults."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            configuration = yaml.safe_load(
                Path(
                    "ppar/setup_templates/axys_apx_analytics/"
                    "axys_apx_analytics.yaml"
                ).read_text(encoding="utf-8")
            )
            self.assertNotIn("confidence_level", configuration["analytics"])
            (site_directory / "ppar.yaml").write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )

            defaulted = _analytics_cli.script_run_settings(site_directory, [])

            settings = _analytics_cli.script_run_settings(
                site_directory,
                ["--confidence-level", "0.90"],
            )

        self.assertEqual(defaulted.confidence_level, 0.95)
        self.assertEqual(
            defaulted.holidays_path,
            site_directory / "holidays.csv",
        )
        self.assertEqual(settings.confidence_level, 0.90)

    def test_analytics_required_values_may_come_from_command_line(self) -> None:
        """Portfolio and benchmark may be supplied by YAML or command line."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            configuration = yaml.safe_load(
                Path(
                    "ppar/setup_templates/axys_apx_analytics/"
                    "axys_apx_analytics.yaml"
                ).read_text(encoding="utf-8")
            )
            del configuration["analytics"]["portfolio"]
            del configuration["analytics"]["benchmark"]
            (site_directory / "ppar.yaml").write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(
                ValueError,
                "analytics.portfolio must be set in ppar.yaml or supplied",
            ):
                _analytics_cli.script_run_settings(site_directory, [])

            settings = _analytics_cli.script_run_settings(
                site_directory,
                ["--portfolio", "CLI_PORT", "--benchmark", "CLI_BENCH"],
            )

        self.assertEqual(settings.portfolio_code, "CLI_PORT")
        self.assertEqual(settings.benchmark_code, "CLI_BENCH")

    def test_analytics_omitted_optional_settings_use_documented_defaults(self) -> None:
        """Every optional Analytics run setting has an executable default."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            configuration = {
                "analytics": {
                    "portfolio": "PORT",
                    "benchmark": "BENCH",
                }
            }
            (site_directory / "ppar.yaml").write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )

            settings = _analytics_cli.script_run_settings(site_directory, [])

        self.assertEqual(
            settings.frequency,
            _analytics_cli.Frequency.AS_OFTEN_AS_POSSIBLE,
        )
        self.assertIsNone(settings.holidays_path)
        self.assertEqual(settings.output_directory, site_directory / "output")
        self.assertIsNone(settings.from_date)
        self.assertIsNone(settings.thru_date)
        self.assertEqual(settings.classification_name, "Security")
        self.assertEqual(settings.annual_minimum_acceptable_return, 0.0)
        self.assertEqual(settings.annual_risk_free_rate, 0.03)
        self.assertEqual(settings.confidence_level, 0.95)
        self.assertEqual(settings.portfolio_value, 100000.0)
        self.assertEqual(settings.currency_symbol, "$")

    def test_analytics_run_settings_reject_unknown_yaml_keys(self) -> None:
        """Analytics setting typos fail instead of being ignored."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            configuration = yaml.safe_load(
                Path(
                    "ppar/setup_templates/axys_apx_analytics/"
                    "axys_apx_analytics.yaml"
                ).read_text(encoding="utf-8")
            )
            configuration["analytics"]["confidence_levle"] = 0.95
            (site_directory / "ppar.yaml").write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(
                PpaError,
                "analytics has unsupported keys: confidence_levle",
            ):
                _analytics_cli.script_run_settings(site_directory, [])

    def test_report_cli_modules_expose_help(self) -> None:
        """Report CLI modules expose consistent command-line help."""
        module_expectations = {
            _SETUP_MODULE: (
                "Create a PPAR Audit workspace"
            ),
            _ANALYTICS_MODULE: (
                "Write Axys/APX analytics reports"
            ),
            _SITE_REPORT_MODULE: (
                "Write PPAR Audit review packages"
            ),
            _VALIDATE_BUNDLE_MODULE: (
                "Validate an Audit report bundle."
            ),
            _VALIDATE_CONFIG_MODULE: (
                "Validate an Audit YAML configuration."
            ),
        }

        for module_name, expected_description in module_expectations.items():
            with self.subTest(module_name=module_name):
                result = subprocess.run(
                    _module_command(module_name, "--help"),
                    check=True,
                    capture_output=True,
                    text=True,
                )

                self.assertIn(expected_description, result.stdout)
                self.assertIn("-h, --help", result.stdout)
                self.assertEqual(result.stderr, "")
                if module_name == _SETUP_MODULE:
                    self.assertIn("--analytics", result.stdout)
                    self.assertIn("--generic-analytics", result.stdout)
                    self.assertIn("--overwrite", result.stdout)
                    self.assertNotIn("--guide", result.stdout)
                    self.assertNotIn("--include-generic-analytics", result.stdout)
                    self.assertIn("usage: ppar setup", result.stdout)
                    self.assertIn("ppar setup ./my_ppar_audit", result.stdout)
                    self.assertIn(
                        "ppar setup ./my_ppar_analytics --analytics",
                        result.stdout,
                    )
                    self.assertIn(
                        "ppar setup ./my_ppar_generic_analytics --generic-analytics",
                        result.stdout,
                    )
                if module_name == _ANALYTICS_MODULE:
                    self.assertIn("usage: ppar analytics", result.stdout)
                    self.assertIn(
                        "ppar analytics ./my_ppar_analytics",
                        result.stdout,
                    )
                if module_name == _SITE_REPORT_MODULE:
                    self.assertNotIn("--report", result.stdout)
                    self.assertIn("usage: ppar audit", result.stdout)
                    self.assertIn("ppar audit ./my_ppar_audit", result.stdout)

    def test_top_level_ppar_cli_keeps_general_help_audit_focused(self) -> None:
        """General help presents the current Audit product and onboarding path."""
        result = subprocess.run(
            _module_command(_PPAR_MODULE, "--help"),
            check=True,
            capture_output=True,
            text=True,
        )

        self.assertIn("usage: ppar <command> [options]", result.stdout)
        self.assertIn("setup", result.stdout)
        self.assertIn("audit", result.stdout)
        self.assertIn("Write Audit reports", result.stdout)
        self.assertNotIn("analytics", result.stdout.lower())
        self.assertNotIn("performance_comparison", result.stdout)
        self.assertNotIn("perfcomp", result.stdout)
        self.assertNotIn("PPAR command-line tools", result.stdout)
        self.assertIn("Examples:", result.stdout)
        self.assertIn("ppar setup ./my_ppar_audit", result.stdout)
        self.assertIn("ppar audit ./my_ppar_audit", result.stdout)
        self.assertNotIn("Set up and run PPAR reports.", result.stdout)
        self.assertNotIn("After setup", result.stdout)
        self.assertEqual(result.stderr, "")

    def test_top_level_ppar_cli_without_args_prints_first_run_handoff(self) -> None:
        """Typing ``ppar`` gives new users the setup command, not parser noise."""
        result = subprocess.run(
            _module_command(_PPAR_MODULE),
            check=True,
            capture_output=True,
            text=True,
        )

        self.assertIn(
            "PPAR Audit explains why reported portfolio performance changed.",
            result.stdout,
        )
        self.assertIn("First-time setup:", result.stdout)
        self.assertIn("ppar setup ./my_ppar_audit", result.stdout)
        self.assertIn("ppar audit ./my_ppar_audit", result.stdout)
        self.assertNotIn("analytics", result.stdout.lower())
        self.assertNotIn("usage:", result.stdout)
        self.assertEqual(result.stderr, "")

    def test_setup_requires_workspace_directory(self) -> None:
        """Setup requires an explicit destination folder."""
        result = subprocess.run(
            _module_command(_SETUP_MODULE),
            check=False,
            capture_output=True,
            text=True,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("usage: ppar setup", result.stderr)
        self.assertIn("workspace_directory", result.stderr)
        self.assertNotIn("--guide", result.stderr)

    def test_setup_rejects_retired_and_conflicting_analytics_modes(self) -> None:
        """Setup exposes only the two current, mutually exclusive opt-in modes."""
        with tempfile.TemporaryDirectory() as directory:
            workspace_directory = Path(directory) / "workspace"
            retired_result = subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(workspace_directory),
                    "--include-generic-analytics",
                ),
                check=False,
                capture_output=True,
                text=True,
            )
            conflicting_result = subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(workspace_directory),
                    "--analytics",
                    "--generic-analytics",
                ),
                check=False,
                capture_output=True,
                text=True,
            )

        self.assertNotEqual(retired_result.returncode, 0)
        self.assertIn(
            "unrecognized arguments: --include-generic-analytics",
            retired_result.stderr,
        )
        self.assertNotEqual(conflicting_result.returncode, 0)
        self.assertIn("not allowed with argument --analytics", conflicting_result.stderr)

    def test_setup_writes_canonical_audit_workspace(self) -> None:
        """Default setup creates one self-describing Audit workspace."""
        with tempfile.TemporaryDirectory() as directory:
            workspace_directory = Path(directory) / "my_ppar_audit"

            result = subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(workspace_directory),
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            config_path = workspace_directory / "ppar.yaml"
            config = yaml.safe_load(config_path.read_text(encoding="utf-8"))
            readme = (workspace_directory / "README.md").read_text(encoding="utf-8")
            self.assertIn("PPAR Audit workspace ready:", result.stdout)
            self.assertIn("To run Audit:", result.stdout)
            self.assertNotIn("Analytics", result.stdout)
            self.assertIn("To customize with your own data:", result.stdout)
            self.assertIn(
                f"Refer to the \"Customizing With Your Own Data\" section in "
                f"{workspace_directory / 'README.md'}",
                result.stdout,
            )
            self.assertNotIn("(created)", result.stdout)
            self.assertNotIn("(written)", result.stdout)
            self.assertTrue((workspace_directory / "README.md").exists())
            self.assertIn("# PPAR Audit Workspace", readme)
            self.assertIn("## First Run", readme)
            self.assertIn("## Customizing With Your Own Data", readme)
            self.assertIn("## Folder Map", readme)
            self.assertNotIn("PYTHON_TUTORIAL.md", readme)
            self.assertNotIn("Performance Analytics", readme)
            self.assertIn("run_audit.py", readme)
            self.assertIn("ppar audit -h", readme)
            self.assertIn("python run_audit.py -h", readme)
            self.assertNotIn("run_portfolio_comparison.py", readme)
            self.assertNotIn("run_security_comparison.py", readme)
            self.assertIn("Audit compares two snapshots", readme)
            self.assertIn("### Getting Data from Axys/APX", readme)
            self.assertIn("Start by reviewing the comments under `files:`", readme)
            self.assertIn(
                "The report labels that outcome as **Fully Explained**",
                readme,
            )
            self.assertIn("use a REP performance or attribution", readme)
            self.assertIn("try IMEX first", readme)
            self.assertIn("source/destination and special-security context", readme)
            self.assertIn("PPAR-normalized examples", readme)
            self.assertEqual(readme.count("fx_rates.csv"), 2)
            self.assertEqual(readme.count("splits.csv"), 2)
            section_sequence = [
                "## What This Folder Is For",
                "## First Run",
                "## Customizing With Your Own Data",
                "## Optional Python Script",
                "## Folder Map",
            ]
            for before, after in zip(section_sequence, section_sequence[1:]):
                with self.subTest(before=before, after=after):
                    self.assertLess(readme.index(before), readme.index(after))
            self.assertIn("the original or older source-data snapshot", readme)
            self.assertIn("the newer, corrected, or restated source-data snapshot", readme)
            self.assertIn("Edit `ppar.yaml`.", readme)
            self.assertTrue((workspace_directory / "run_audit.py").exists())
            audit_script = (workspace_directory / "run_audit.py").read_text(
                encoding="utf-8"
            )
            self.assertIn("from ppar.audit.cli.site_report import run_report", audit_script)
            self.assertIn("result = run_report(", audit_script)
            self.assertIn("Optional one-run customization examples", audit_script)
            self.assertNotIn("site_report.main(", audit_script)
            self.assertFalse(
                (workspace_directory / "run_portfolio_comparison.py").exists()
            )
            self.assertFalse(
                (workspace_directory / "run_security_comparison.py").exists()
            )
            self.assertFalse((workspace_directory / "PYTHON_TUTORIAL.md").exists())
            self.assertFalse((workspace_directory / "generic_analytics").exists())
            self.assertFalse((workspace_directory / "analytics").exists())
            self.assertFalse((workspace_directory / "audit").exists())
            self.assertEqual(config["snapshots"]["a"]["path"], "snapshot_a")
            self.assertEqual(config["snapshots"]["b"]["path"], "snapshot_b")
            for snapshot_name in ("a", "b"):
                snapshot = config["snapshots"][snapshot_name]
                self.assertNotIn("vendor", snapshot)
                self.assertNotIn("schema", snapshot)
            for file_name in (
                "portfolio_performance",
                "security_performance",
                "security_master",
                "holdings",
                "transactions",
                "splits",
                "fx_rates",
            ):
                self.assertIn("columns", config["files"][file_name])
            self.assertNotIn("security_id", config)
            self.assertIn("security_return_reconstruction", config)
            self.assertFalse((workspace_directory / "column_mappings.yaml").exists())
            for snapshot_name in ("snapshot_a", "snapshot_b"):
                self.assertTrue(
                    (workspace_directory / snapshot_name / "secmast.csv").exists()
                )

    def test_setup_can_create_analytics_workspace(self) -> None:
        """The explicit Analytics mode creates one Analytics workspace."""
        with tempfile.TemporaryDirectory() as directory:
            workspace_directory = Path(directory) / "my_ppar_analytics"

            result = subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(workspace_directory),
                    "--analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("PPAR Analytics workspace ready:", result.stdout)
            self.assertIn(f"ppar analytics {workspace_directory}", result.stdout)
            self.assertNotIn("ppar audit", result.stdout)
            self.assertNotIn("secperf.csv", result.stdout)
            readme = (workspace_directory / "README.md").read_text(encoding="utf-8")
            self.assertIn("# PPAR Analytics Workspace", readme)
            self.assertIn('pip install "ppar[analytics]"', readme)
            self.assertIn(
                f"ppar setup {workspace_directory} --analytics",
                readme,
            )
            self.assertNotIn("PPAR Audit", readme)
            self.assertTrue((workspace_directory / "run_analytics.py").exists())
            self.assertTrue((workspace_directory / "ppar.yaml").exists())
            self.assertTrue((workspace_directory / "portperf.csv").exists())
            self.assertTrue((workspace_directory / "secperf.csv").exists())
            self.assertTrue((workspace_directory / "secmast.csv").exists())
            self.assertEqual(
                (workspace_directory / "holidays.csv").read_text(encoding="utf-8"),
                "2024-03-29\n",
            )
            self.assertFalse((workspace_directory / "snapshot_a").exists())
            self.assertFalse((workspace_directory / "audit").exists())
            analytics_script = (
                workspace_directory / "run_analytics.py"
            ).read_text(encoding="utf-8")
            self.assertNotIn("ppar.demos", analytics_script)
            self.assertNotIn("TemporaryDirectory", analytics_script)
            self.assertNotIn("MPLCONFIGDIR", analytics_script)
            self.assertIn("SPECIFICATIONS_PATH", analytics_script)
            self.assertIn("AxysData", analytics_script)
            self.assertIn("to_analytics", analytics_script)

    def test_setup_rerun_preserves_user_edits_without_overwrite(self) -> None:
        """Setup does not replace local user edits unless overwrite is requested."""
        with tempfile.TemporaryDirectory() as directory:
            workspace_directory = Path(directory) / "my_ppar_audit"
            subprocess.run(
                _module_command(_SETUP_MODULE, str(workspace_directory)),
                check=True,
                capture_output=True,
                text=True,
            )
            readme_path = workspace_directory / "README.md"
            audit_script_path = workspace_directory / "run_audit.py"
            audit_config_path = workspace_directory / "ppar.yaml"

            custom_readme = "custom readme\n"
            custom_audit_script = "# custom audit script\n"
            custom_comparison_config = (
                audit_config_path.read_text(encoding="utf-8")
                + "\n# custom performance comparison note\n"
            )
            readme_path.write_text(custom_readme, encoding="utf-8")
            audit_script_path.write_text(custom_audit_script, encoding="utf-8")
            audit_config_path.write_text(
                custom_comparison_config,
                encoding="utf-8",
            )

            subprocess.run(
                _module_command(_SETUP_MODULE, str(workspace_directory)),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertEqual(readme_path.read_text(encoding="utf-8"), custom_readme)
            self.assertEqual(
                audit_script_path.read_text(encoding="utf-8"),
                custom_audit_script,
            )
            self.assertEqual(
                audit_config_path.read_text(encoding="utf-8"),
                custom_comparison_config,
            )

    def test_setup_rejects_mixed_or_legacy_workspace_roots(self) -> None:
        """Setup refuses to mix workflow files into one workspace."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            audit_workspace = root / "my_ppar_audit"
            subprocess.run(
                _module_command(_SETUP_MODULE, str(audit_workspace)),
                check=True,
                capture_output=True,
                text=True,
            )

            mixed_result = subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(audit_workspace),
                    "--analytics",
                ),
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(mixed_result.returncode, 0)
            self.assertIn(
                "contains an existing audit workspace",
                mixed_result.stderr,
            )
            legacy_root = root / "legacy"
            (legacy_root / "audit").mkdir(parents=True)
            (legacy_root / "audit" / "ppar.yaml").write_text(
                "audit: {}\n",
                encoding="utf-8",
            )
            legacy_result = subprocess.run(
                _module_command(_SETUP_MODULE, str(legacy_root)),
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(legacy_result.returncode, 0)
            self.assertIn("legacy combined PPAR workspace", legacy_result.stderr)

    def test_setup_can_create_generic_analytics_workspace(self) -> None:
        """The explicit Generic Analytics mode creates a standalone workspace."""
        with tempfile.TemporaryDirectory() as directory:
            workspace_directory = Path(directory) / "my_ppar_generic_analytics"

            result = subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(workspace_directory),
                    "--generic-analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            generic_directory = workspace_directory
            self.assertIn("PPAR Generic Analytics workspace ready:", result.stdout)
            self.assertIn("To run Generic Analytics:", result.stdout)
            self.assertIn(
                f"python {generic_directory / 'run_generic_analytics.py'}",
                result.stdout,
            )
            self.assertTrue((generic_directory / "README.md").exists())
            readme = (generic_directory / "README.md").read_text(encoding="utf-8")
            self.assertIn("# PPAR Generic Analytics Workspace", readme)
            self.assertIn("## Customizing With Your Own Data", readme)
            self.assertIn("There is no `ppar.yaml`", readme)
            self.assertFalse((generic_directory / "ppar.yaml").exists())
            self.assertTrue(
                (generic_directory / "run_generic_analytics.py").exists()
            )
            self.assertEqual(
                (generic_directory / "holidays.csv").read_text(encoding="utf-8"),
                "2024-03-29\n",
            )
            generic_script = (
                generic_directory / "run_generic_analytics.py"
            ).read_text(encoding="utf-8")
            self.assertNotIn("ppar.demos", generic_script)
            self.assertNotIn("TemporaryDirectory", generic_script)
            self.assertNotIn("MPLCONFIGDIR", generic_script)
            self.assertTrue(
                (
                    generic_directory
                    / "performance"
                    / "Mega-Cap Alpha Portfolio.csv"
                ).exists()
            )
            self.assertTrue(
                (
                    generic_directory
                    / "classifications"
                    / "Economic Sector.csv"
                ).exists()
            )

    def test_setup_rejects_mixing_generic_and_configured_workspaces(self) -> None:
        """Generic Analytics cannot be combined with Audit or Analytics files."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            generic_directory = root / "my_ppar_generic_analytics"
            audit_directory = root / "my_ppar_audit"
            subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(generic_directory),
                    "--generic-analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )
            subprocess.run(
                _module_command(_SETUP_MODULE, str(audit_directory)),
                check=True,
                capture_output=True,
                text=True,
            )

            audit_into_generic = subprocess.run(
                _module_command(_SETUP_MODULE, str(generic_directory)),
                check=False,
                capture_output=True,
                text=True,
            )
            generic_into_audit = subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(audit_directory),
                    "--generic-analytics",
                ),
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(audit_into_generic.returncode, 0)
            self.assertIn(
                "contains an existing generic_analytics workspace",
                audit_into_generic.stderr,
            )
            self.assertNotEqual(generic_into_audit.returncode, 0)
            self.assertIn(
                "contains an existing audit workspace",
                generic_into_audit.stderr,
            )

    def test_setup_installed_python_scripts_run_end_to_end(self) -> None:
        """Copied setup scripts are the canonical Python smoke-test path."""
        with tempfile.TemporaryDirectory() as directory:
            audit_directory = Path(directory) / "my_ppar_audit"
            analytics_directory = Path(directory) / "my_ppar_analytics"
            generic_directory = Path(directory) / "my_ppar_generic_analytics"
            subprocess.run(
                _module_command(
                    _PPAR_MODULE,
                    "setup",
                    str(audit_directory),
                ),
                check=True,
                capture_output=True,
                text=True,
            )
            subprocess.run(
                _module_command(
                    _PPAR_MODULE,
                    "setup",
                    str(analytics_directory),
                    "--analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )
            subprocess.run(
                _module_command(
                    _PPAR_MODULE,
                    "setup",
                    str(generic_directory),
                    "--generic-analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            script_paths = (
                audit_directory / "run_audit.py",
                analytics_directory / "run_analytics.py",
                generic_directory / "run_generic_analytics.py",
            )
            for script_path in script_paths:
                with self.subTest(script_path=script_path.name):
                    result = subprocess.run(
                        [sys.executable, str(script_path)],
                        check=True,
                        capture_output=True,
                        text=True,
                    )
                    self.assertIn("Open these files to review", result.stdout)
                    self.assertEqual(result.stderr, "")

            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.xlsx"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "security"
                    / "security_audit.xlsx"
                ).exists()
            )
            self.assertTrue(
                (
                    analytics_directory
                    / "output"
                    / "risk_statistics.html"
                ).exists()
            )
            self.assertTrue(
                (
                    generic_directory
                    / "output"
                    / "risk_statistics.html"
                ).exists()
            )

    def test_setup_audit_script_matches_default_cli_workflow(self) -> None:
        """The visible Python example stays equivalent to ``ppar audit``."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            cli_site = root / "cli_site"
            script_site = root / "script_site"
            for site in (cli_site, script_site):
                subprocess.run(
                    _module_command(_PPAR_MODULE, "setup", str(site)),
                    check=True,
                    capture_output=True,
                    text=True,
                )

            cli_audit = cli_site
            script_audit = script_site
            subprocess.run(
                _module_command(_PPAR_MODULE, "audit", str(cli_audit)),
                check=True,
                capture_output=True,
                text=True,
            )
            subprocess.run(
                [sys.executable, str(script_audit / "run_audit.py")],
                check=True,
                capture_output=True,
                text=True,
            )

            cli_output = cli_audit / "output"
            script_output = script_audit / "output"
            cli_files = {
                path.relative_to(cli_output)
                for path in cli_output.rglob("*")
                if path.is_file()
            }
            script_files = {
                path.relative_to(script_output)
                for path in script_output.rglob("*")
                if path.is_file()
            }
            self.assertEqual(cli_files, script_files)

            for report_level in ("portfolio", "security"):
                with self.subTest(report_level=report_level):
                    audit_stem = f"{report_level}_audit"
                    relative_html = Path(report_level) / f"{audit_stem}.html"
                    self.assertEqual(
                        (cli_output / relative_html).read_text(encoding="utf-8"),
                        (script_output / relative_html).read_text(encoding="utf-8"),
                    )
                    cli_workbook = load_workbook(
                        cli_output / report_level / f"{audit_stem}.xlsx",
                        read_only=True,
                        data_only=False,
                    )
                    script_workbook = load_workbook(
                        script_output / report_level / f"{audit_stem}.xlsx",
                        read_only=True,
                        data_only=False,
                    )
                    self.assertEqual(cli_workbook.sheetnames, script_workbook.sheetnames)
                    for sheet_name in cli_workbook.sheetnames:
                        self.assertEqual(
                            list(cli_workbook[sheet_name].values),
                            list(script_workbook[sheet_name].values),
                        )
                    cli_workbook.close()
                    script_workbook.close()

            audit_help = subprocess.run(
                [sys.executable, str(script_audit / "run_audit.py"), "--help"],
                check=True,
                capture_output=True,
                text=True,
            ).stdout
            self.assertIn("Run Audit through Python", audit_help)
            self.assertIn("Edit the run_report() call", audit_help)
            self.assertNotIn("--output-directory", audit_help)
            self.assertNotIn("--title", audit_help)
            self.assertNotIn("--html-only", audit_help)

    def test_setup_analytics_script_matches_default_cli_workflow(self) -> None:
        """The visible Python example stays equivalent to ``ppar analytics``."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            cli_site = root / "cli_site"
            script_site = root / "script_site"
            for site in (cli_site, script_site):
                subprocess.run(
                    _module_command(
                        _PPAR_MODULE,
                        "setup",
                        str(site),
                        "--analytics",
                    ),
                    check=True,
                    capture_output=True,
                    text=True,
                )

            cli_analytics = cli_site
            script_analytics = script_site
            subprocess.run(
                _module_command(_PPAR_MODULE, "analytics", str(cli_analytics)),
                check=True,
                capture_output=True,
                text=True,
            )
            subprocess.run(
                [sys.executable, str(script_analytics / "run_analytics.py")],
                check=True,
                capture_output=True,
                text=True,
            )

            cli_output = cli_analytics / "output"
            script_output = script_analytics / "output"
            cli_files = {
                path.relative_to(cli_output)
                for path in cli_output.rglob("*")
                if path.is_file()
            }
            script_files = {
                path.relative_to(script_output)
                for path in script_output.rglob("*")
                if path.is_file()
            }
            self.assertEqual(cli_files, script_files)
            for relative_path in cli_files:
                with self.subTest(relative_path=relative_path.as_posix()):
                    self.assertEqual(
                        (cli_output / relative_path).read_bytes(),
                        (script_output / relative_path).read_bytes(),
                    )

            cli_override_output = root / "cli_override_output"
            script_override_output = root / "script_override_output"
            subprocess.run(
                _module_command(
                    _PPAR_MODULE,
                    "analytics",
                    str(cli_analytics),
                    "--frequency",
                    "yearly",
                    "--from-date",
                    "2022-01-01",
                    "--thru-date",
                    "2025-12-31",
                    "--classification",
                    "Security",
                    "--annual-minimum-acceptable-return",
                    "0.02",
                    "--annual-risk-free-rate",
                    "0.04",
                    "--confidence-level",
                    "0.90",
                    "--portfolio-value",
                    "250000",
                    "--currency-symbol",
                    "EUR",
                    "--output-directory",
                    str(cli_override_output),
                ),
                check=True,
                capture_output=True,
                text=True,
            )
            subprocess.run(
                [
                    sys.executable,
                    str(script_analytics / "run_analytics.py"),
                    "--frequency",
                    "yearly",
                    "--from-date",
                    "2022-01-01",
                    "--thru-date",
                    "2025-12-31",
                    "--classification",
                    "Security",
                    "--annual-minimum-acceptable-return",
                    "0.02",
                    "--annual-risk-free-rate",
                    "0.04",
                    "--confidence-level",
                    "0.90",
                    "--portfolio-value",
                    "250000",
                    "--currency-symbol",
                    "EUR",
                    "--output-directory",
                    str(script_override_output),
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            override_files = {
                path.relative_to(cli_override_output)
                for path in cli_override_output.rglob("*")
                if path.is_file()
            }
            self.assertEqual(
                override_files,
                {
                    path.relative_to(script_override_output)
                    for path in script_override_output.rglob("*")
                    if path.is_file()
                },
            )
            for relative_path in override_files:
                self.assertEqual(
                    (cli_override_output / relative_path).read_bytes(),
                    (script_override_output / relative_path).read_bytes(),
                )

            analytics_help = subprocess.run(
                [sys.executable, str(script_analytics / "run_analytics.py"), "--help"],
                check=True,
                capture_output=True,
                text=True,
            ).stdout
            for option in (
                "--portfolio",
                "--benchmark",
                "--frequency",
                "--holidays",
                "--output-directory",
                "--from-date",
                "--thru-date",
                "--classification",
                "--annual-minimum-acceptable-return",
                "--annual-risk-free-rate",
                "--confidence-level",
                "--portfolio-value",
                "--currency-symbol",
            ):
                self.assertIn(option, analytics_help)
            for yaml_setting in (
                "YAML analytics.portfolio for this run",
                "YAML analytics.benchmark for this run",
                "YAML analytics.frequency for this run",
                "YAML analytics.holidays",
                "YAML analytics.output_directory for this run",
                "YAML analytics.from_date",
                "YAML analytics.thru_date",
                "YAML analytics.classification for this run",
                "YAML analytics.annual_minimum_acceptable_return for this run",
                "YAML analytics.annual_risk_free_rate for this run",
                "YAML analytics.confidence_level for this run",
                "YAML analytics.portfolio_value for this run",
                "YAML analytics.currency_symbol for this run",
            ):
                self.assertIn(yaml_setting, analytics_help)

    def test_audit_commands_reject_removed_report_option(self) -> None:
        """Neither Audit entrypoint restores report selection."""
        with tempfile.TemporaryDirectory() as directory:
            site = Path(directory) / "site"
            subprocess.run(
                _module_command(_PPAR_MODULE, "setup", str(site)),
                check=True,
                capture_output=True,
                text=True,
            )
            audit_directory = site
            commands = (
                _module_command(
                    _PPAR_MODULE,
                    "audit",
                    str(audit_directory),
                    "--report",
                    "portfolio",
                ),
                [
                    sys.executable,
                    str(audit_directory / "run_audit.py"),
                    "--report",
                    "portfolio",
                ],
            )
            for command in commands:
                with self.subTest(command=command):
                    result = subprocess.run(
                        command,
                        check=False,
                        capture_output=True,
                        text=True,
                    )
                    self.assertEqual(result.returncode, 2)
                    self.assertIn(
                        "unrecognized arguments: --report portfolio",
                        result.stderr,
                    )

            audit_help = subprocess.run(
                [sys.executable, str(audit_directory / "run_audit.py"), "--help"],
                check=True,
                capture_output=True,
                text=True,
            ).stdout
            self.assertNotIn("--report", audit_help)
            self.assertIn("Edit the run_report() call", audit_help)
            self.assertNotIn("--html-only", audit_help)
            self.assertNotIn("--xlsx-only", audit_help)
            self.assertNotIn("--csv-only", audit_help)
            self.assertNotIn("--expand-supporting-files", audit_help)
            self.assertNotIn("--exclude-suppressed", audit_help)
            self.assertNotIn("--allow-incomplete-yaml", audit_help)
            self.assertNotIn("{portfolio,security,both}", audit_help)

    def test_public_python_entrypoints_accept_string_site_directories(self) -> None:
        """Programmatic entrypoints accept string paths as well as ``Path`` values."""
        from ppar.analytics.cli import run_analytics
        from ppar.errors import PpaError
        from ppar.audit.cli.site_report import run_report

        with tempfile.TemporaryDirectory() as directory:
            missing_analytics_path = str(Path(directory) / "missing_analytics")
            missing_comparison_path = str(Path(directory) / "missing_comparison")

            with self.assertRaises(PpaError):
                run_analytics(missing_analytics_path)
            with self.assertRaises(PpaError):
                run_report(missing_comparison_path)

    def test_site_report_writes_both_reports_by_default(self) -> None:
        """The production comparison command writes both workbooks by default."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "my_ppar_audit"
            subprocess.run(
                _module_command(_SETUP_MODULE, str(site_directory)),
                check=True,
                capture_output=True,
                text=True,
            )
            audit_directory = site_directory

            result = subprocess.run(
                _module_command(
                    _SITE_REPORT_MODULE,
                    str(audit_directory),
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn(
                "Open these files to review Audit output:",
                result.stdout,
            )
            self.assertIn("output/portfolio/portfolio_audit.xlsx", result.stdout)
            self.assertIn("output/security/security_audit.xlsx", result.stdout)
            self.assertIn("output/portfolio/portfolio_audit.html", result.stdout)
            self.assertIn("output/security/security_audit.html", result.stdout)
            self.assertTrue((audit_directory / "ppar.yaml").exists())
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.xlsx"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.html"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "security"
                    / "security_audit.xlsx"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "security"
                    / "security_audit.html"
                ).exists()
            )

    def test_site_report_can_write_xlsx_only_output(self) -> None:
        """The production report command can write an XLSX-only audit."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "my_ppar_audit"
            subprocess.run(
                _module_command(_SETUP_MODULE, str(site_directory)),
                check=True,
                capture_output=True,
                text=True,
            )
            audit_directory = site_directory

            subprocess.run(
                _module_command(
                    _SITE_REPORT_MODULE,
                    str(audit_directory),
                    "--xlsx-only",
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            output_directory = audit_directory / "output" / "portfolio"
            self.assertTrue((output_directory / "portfolio_audit.xlsx").exists())
            self.assertFalse((output_directory / "portfolio_audit.html").exists())
            security_output = audit_directory / "output" / "security"
            self.assertTrue((security_output / "security_audit.xlsx").exists())
            self.assertFalse((security_output / "security_audit.html").exists())

    def test_site_report_can_write_csv_only_output(self) -> None:
        """CSV-only mode promotes the canonical CSV review files."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "my_ppar_audit"
            subprocess.run(
                _module_command(_SETUP_MODULE, str(site_directory)),
                check=True,
                capture_output=True,
                text=True,
            )
            audit_directory = site_directory

            result = subprocess.run(
                _module_command(
                    _SITE_REPORT_MODULE,
                    str(audit_directory),
                    "--csv-only",
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            for report_level in ("portfolio", "security"):
                output_directory = audit_directory / "output" / report_level
                self.assertFalse(
                    (output_directory / f"{report_level}_audit.xlsx").exists()
                )
                self.assertFalse(
                    (output_directory / f"{report_level}_audit.html").exists()
                )
                for file_name in (
                    "performance_differences.csv",
                    "performance_difference_causes.csv",
                    "data_issues.csv",
                    "source_detail.csv",
                ):
                    with self.subTest(
                        report_level=report_level,
                        file_name=file_name,
                    ):
                        self.assertTrue((output_directory / file_name).is_file())
                self.assertTrue((output_directory / "audit_support.zip").is_file())
            self.assertIn("performance_differences.csv", result.stdout)
            self.assertIn("performance_difference_causes.csv", result.stdout)
            self.assertIn("data_issues.csv", result.stdout)
            self.assertNotIn("source_detail.csv", result.stdout)

    def test_site_report_shares_reconstruction_cache_between_views(self) -> None:
        """One Audit run reuses reconstruction inputs for both report levels."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            _write_audit_run_settings(site_directory / "ppar.yaml")
            comparison_views = mock.Mock()
            comparison_views.findings.side_effect = [
                mock.sentinel.portfolio_findings,
                mock.sentinel.security_findings,
            ]
            with (
                mock.patch.object(
                    _site_report,
                    "AuditComparisonViews",
                    return_value=comparison_views,
                ),
                mock.patch.object(
                    _site_report._data_issue_checks,
                    "data_issues_table",
                    return_value=mock.sentinel.data_issues,
                ),
                mock.patch.object(
                    _site_report,
                    "_write_report_bundle",
                    side_effect=lambda _config, _findings, output, **kwargs: [
                        output
                        / (
                            f"{kwargs['comparison_level']}_audit.xlsx"
                        )
                    ],
                ) as write_report_bundle,
            ):
                _site_report.run_report(site_directory)

        portfolio_cache = write_report_bundle.call_args_list[0].kwargs[
            "_reconstruction_cache"
        ]
        security_cache = write_report_bundle.call_args_list[1].kwargs[
            "_reconstruction_cache"
        ]
        self.assertIs(portfolio_cache, security_cache)

    def test_site_report_fails_closed_for_other_security_errors(self) -> None:
        """Only unavailable security-performance input permits a security skip."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            _write_audit_run_settings(site_directory / "ppar.yaml")
            comparison_views = mock.Mock()
            comparison_views.findings.side_effect = (
                mock.sentinel.portfolio_findings,
                PpaError("malformed security data", None),
            )
            with (
                mock.patch.object(
                    _site_report,
                    "AuditComparisonViews",
                    return_value=comparison_views,
                ),
                mock.patch.object(
                    _site_report._data_issue_checks,
                    "data_issues_table",
                    return_value=mock.sentinel.data_issues,
                ),
                mock.patch.object(
                    _site_report,
                    "_write_report_bundle",
                    return_value=[site_directory / "portfolio_audit.xlsx"],
                ) as write_report_bundle,
            ):
                with self.assertRaisesRegex(PpaError, "malformed security data"):
                    _site_report.run_report(site_directory)

        write_report_bundle.assert_not_called()

    def test_site_report_write_failure_preserves_both_previous_views(self) -> None:
        """A late security write failure does not promote the portfolio view."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory)
            _write_audit_run_settings(site_directory / "ppar.yaml")
            output_root = site_directory / "output"
            for level in ("portfolio", "security"):
                level_directory = output_root / level
                level_directory.mkdir(parents=True)
                (level_directory / "sentinel.txt").write_text(
                    f"previous {level}",
                    encoding="utf-8",
                )

            comparison_views = mock.Mock()
            comparison_views.findings.side_effect = [
                mock.sentinel.portfolio_findings,
                mock.sentinel.security_findings,
            ]

            def staged_write(
                _config: Path,
                _findings: object,
                output_directory: Path,
                *,
                comparison_level: str,
                **_kwargs: object,
            ) -> list[Path]:
                if comparison_level == "security":
                    raise PpaError("late security write failure", None)
                output_directory.mkdir(parents=True)
                review_path = output_directory / "portfolio_audit.xlsx"
                review_path.write_text("candidate", encoding="utf-8")
                return [review_path]

            with (
                mock.patch.object(
                    _site_report,
                    "AuditComparisonViews",
                    return_value=comparison_views,
                ),
                mock.patch.object(
                    _site_report._data_issue_checks,
                    "data_issues_table",
                    return_value=mock.sentinel.data_issues,
                ),
                mock.patch.object(
                    _site_report,
                    "_write_report_bundle",
                    side_effect=staged_write,
                ),
                self.assertRaisesRegex(PpaError, "late security write failure"),
            ):
                _site_report.run_report(site_directory)

            self.assertEqual(
                (output_root / "portfolio" / "sentinel.txt").read_text(
                    encoding="utf-8"
                ),
                "previous portfolio",
            )
            self.assertEqual(
                (output_root / "security" / "sentinel.txt").read_text(
                    encoding="utf-8"
                ),
                "previous security",
            )
            self.assertFalse(
                (output_root / "portfolio" / "portfolio_audit.xlsx").exists()
            )

    def test_audit_skips_unavailable_security_performance(self) -> None:
        """The standard run writes portfolio output and skips unavailable security."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "my_ppar_audit"
            setup_result = subprocess.run(
                _module_command(_SETUP_MODULE, str(site_directory)),
                check=True,
                capture_output=True,
                text=True,
            )
            audit_directory = site_directory
            stale_security_directory = audit_directory / "output" / "security"
            stale_security_directory.mkdir(parents=True)
            (stale_security_directory / "stale.txt").write_text(
                "older run",
                encoding="utf-8",
            )
            (audit_directory / "snapshot_a" / "secperf.csv").unlink()
            (audit_directory / "snapshot_b" / "secperf.csv").unlink()

            default_result = subprocess.run(
                _module_command(_SITE_REPORT_MODULE, str(audit_directory)),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("PPAR Audit workspace ready:", setup_result.stdout)
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.xlsx"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.html"
                ).exists()
            )
            self.assertIn(
                "output/portfolio/portfolio_audit.xlsx",
                default_result.stdout,
            )
            self.assertIn(
                "Security output skipped because files.security_performance is not available.",
                default_result.stdout,
            )
            self.assertFalse((audit_directory / "output" / "security").exists())

    def test_site_report_writes_both_available_reports(self) -> None:
        """The production report command writes portfolio and security output."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "my_ppar_audit"
            subprocess.run(
                _module_command(_SETUP_MODULE, str(site_directory)),
                check=True,
                capture_output=True,
                text=True,
            )
            audit_directory = site_directory

            result = subprocess.run(
                _module_command(_SITE_REPORT_MODULE, str(audit_directory)),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn(
                "Open these files to review Audit output:",
                result.stdout,
            )
            self.assertIn("output/security/security_audit.xlsx", result.stdout)
            self.assertIn("output/security/security_audit.html", result.stdout)
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "security"
                    / "security_audit.xlsx"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "security"
                    / "security_audit.html"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.xlsx"
                ).exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.html"
                ).exists()
            )

    def test_analytics_cli_writes_site_outputs(self) -> None:
        """The production analytics command writes output from setup data."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "my_ppar_analytics"
            subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(site_directory),
                    "--analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )
            analytics_directory = site_directory

            result = subprocess.run(
                _module_command(_ANALYTICS_MODULE, str(analytics_directory)),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("Open these files to review analytics output:", result.stdout)
            self.assertIn("risk_statistics.html", result.stdout)
            self.assertNotIn("Using quarterly reporting.", result.stdout)
            self.assertNotIn("Time:", result.stdout)
            self.assertNotIn("Analytics output:", result.stdout)
            self.assertTrue(
                (analytics_directory / "output" / "risk_statistics.html").exists()
            )
            self.assertTrue(
                (
                    analytics_directory
                    / "output"
                    / "sector_overall_attribution.html"
                ).exists()
            )

            self.assertFalse((analytics_directory / "output" / ".matplotlib").exists())
            self.assertFalse((analytics_directory / "output" / ".cache").exists())

    def test_analytics_native_frequency_omits_risk_statistics(self) -> None:
        """Native-period Attribution succeeds without a Risk Statistics artifact."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "my_ppar_analytics"
            subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(site_directory),
                    "--analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )
            analytics_directory = site_directory
            config_path = analytics_directory / "ppar.yaml"
            configuration = yaml.safe_load(config_path.read_text(encoding="utf-8"))
            del configuration["analytics"]["frequency"]
            config_path.write_text(
                yaml.safe_dump(configuration, sort_keys=False),
                encoding="utf-8",
            )

            result = subprocess.run(
                _module_command(_ANALYTICS_MODULE, str(analytics_directory)),
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertIn("Open these files to review analytics output:", result.stdout)
            self.assertNotIn("risk_statistics.html", result.stdout)
            self.assertFalse(
                (analytics_directory / "output" / "risk_statistics.html").exists()
            )
            self.assertTrue(
                (
                    analytics_directory
                    / "output"
                    / "sector_overall_attribution.html"
                ).exists()
            )

            script_output = Path(directory) / "script_output"
            script_result = subprocess.run(
                [
                    sys.executable,
                    str(analytics_directory / "run_analytics.py"),
                    "--output-directory",
                    str(script_output),
                ],
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertEqual(script_result.returncode, 0, script_result.stderr)
            self.assertNotIn("risk_statistics.html", script_result.stdout)
            self.assertFalse((script_output / "risk_statistics.html").exists())
            self.assertTrue(
                (script_output / "sector_overall_attribution.html").exists()
            )

    def test_performance_comparison_demo_handoff_matches_quiet_success_contract(
        self,
    ) -> None:
        """Performance-comparison demo handoffs only list the workbook path."""
        from ppar.audit.cli.site_report import (
            _print_success,
        )

        stdout = io.StringIO()
        with redirect_stdout(stdout):
            _print_success(
                {
                    "review_paths": [
                        Path("_demo_output")
                        / "audit"
                        / "portfolio"
                        / "portfolio_audit.xlsx",
                    ],
                }
            )

        output = stdout.getvalue()
        self.assertIn("Open these files to review Audit output:", output)
        self.assertIn("portfolio_audit.xlsx", output)
        self.assertNotIn("portfolio_audit.html", output)
        self.assertNotIn("manifest.json", output)
        for phrase in _DEMO_QUIET_PHRASES:
            self.assertNotIn(phrase, output)

    def test_top_level_commands_do_not_default_from_unconfigured_root(self) -> None:
        """Production commands require a configured workspace or explicit path."""
        with tempfile.TemporaryDirectory() as directory:
            site_directory = Path(directory) / "unconfigured"
            site_directory.mkdir()

            analytics_result = subprocess.run(
                _module_command(_PPAR_MODULE, "analytics"),
                cwd=site_directory,
                capture_output=True,
                text=True,
            )
            comparison_result = subprocess.run(
                _module_command(_PPAR_MODULE, "audit"),
                cwd=site_directory,
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(analytics_result.returncode, 0)
            self.assertNotEqual(comparison_result.returncode, 0)
            self.assertIn("Analytics failed:", analytics_result.stderr)
            self.assertIn(
                "Run from the Analytics workspace or pass its folder.",
                analytics_result.stderr,
            )
            self.assertIn(
                "ppar setup ./my_ppar_analytics --analytics",
                analytics_result.stderr,
            )
            self.assertIn("Report failed:", comparison_result.stderr)
            self.assertIn(
                "Run from the Audit workspace or pass its folder.",
                comparison_result.stderr,
            )
            self.assertIn("ppar setup ./my_ppar_audit", comparison_result.stderr)

    def test_top_level_commands_default_inside_workflow_folders(self) -> None:
        """Production commands can default to cwd inside their configured folder."""
        with tempfile.TemporaryDirectory() as directory:
            audit_directory = Path(directory) / "my_ppar_audit"
            analytics_directory = Path(directory) / "my_ppar_analytics"
            subprocess.run(
                _module_command(_PPAR_MODULE, "setup", str(audit_directory)),
                check=True,
                capture_output=True,
                text=True,
            )
            subprocess.run(
                _module_command(
                    _PPAR_MODULE,
                    "setup",
                    str(analytics_directory),
                    "--analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            analytics_result = subprocess.run(
                _module_command(_PPAR_MODULE, "analytics"),
                cwd=analytics_directory,
                check=True,
                capture_output=True,
                text=True,
            )
            comparison_result = subprocess.run(
                _module_command(_PPAR_MODULE, "audit"),
                cwd=audit_directory,
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn(
                "Open these files to review analytics output:",
                analytics_result.stdout,
            )
            self.assertIn(
                "Open these files to review Audit output:",
                comparison_result.stdout,
            )
            self.assertTrue(
                (analytics_directory / "output" / "risk_statistics.html").exists()
            )
            self.assertTrue(
                (
                    audit_directory
                    / "output"
                    / "portfolio"
                    / "portfolio_audit.xlsx"
                ).exists()
            )

    def test_analytics_cli_resolves_relative_site_directory_once(self) -> None:
        """Analytics source paths stay config-relative when the site path is relative."""
        with tempfile.TemporaryDirectory(prefix="ppar_relative_site_") as directory:
            site_directory = Path(directory) / "my_ppar_analytics"
            relative_site_directory = Path(os.path.relpath(site_directory, Path.cwd()))
            subprocess.run(
                _module_command(
                    _SETUP_MODULE,
                    str(relative_site_directory),
                    "--analytics",
                ),
                check=True,
                capture_output=True,
                text=True,
            )
            analytics_directory = relative_site_directory

            result = subprocess.run(
                _module_command(_ANALYTICS_MODULE, str(analytics_directory)),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("Open these files to review analytics output:", result.stdout)
            self.assertNotIn("analytics/../", result.stderr)
            self.assertTrue(
                (
                    site_directory
                    / "output"
                    / "risk_statistics.html"
                ).exists()
            )

    def test_validate_bundle_cli_module_accepts_valid_bundle(self) -> None:
        """The bundle validator CLI module accepts a generated bundle."""
        with tempfile.TemporaryDirectory() as directory:
            output_directory = Path(directory) / "bundle"
            self._write_bundle(output_directory)

            result = subprocess.run(
                _module_command(
                    _VALIDATE_BUNDLE_MODULE,
                    str(output_directory),
                ),
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("Bundle validation passed:", result.stdout)
            self.assertIn(str(output_directory), result.stdout)
            self.assertEqual(result.stderr, "")

    def test_validate_bundle_cli_module_reports_invalid_bundle(self) -> None:
        """The bundle validator CLI module exits nonzero for a broken bundle."""
        with tempfile.TemporaryDirectory() as directory:
            output_directory = Path(directory) / "bundle"
            self._write_bundle(output_directory)
            top_evidence_path = (
                output_directory / "supporting_files" / "top_evidence.csv"
            )
            header = top_evidence_path.read_text(encoding="utf-8").splitlines()[0]
            top_evidence_path.write_text(header + "\n", encoding="utf-8")

            result = subprocess.run(
                _module_command(
                    _VALIDATE_BUNDLE_MODULE,
                    str(output_directory),
                ),
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 1)
            self.assertEqual(result.stdout, "")
            self.assertIn("Bundle validation failed:", result.stderr)
        self.assertIn("table 'top_evidence' row count is 0, expected 9", result.stderr)

    def test_validate_config_cli_module_accepts_valid_yaml(self) -> None:
        """The CLI config validator accepts a valid comparison YAML file."""
        result = subprocess.run(
            _module_command(
                _VALIDATE_CONFIG_MODULE,
                str(_PORTFOLIO_COMPARISON_PATH),
            ),
            check=True,
            capture_output=True,
            text=True,
        )

        self.assertIn("Config validation passed:", result.stdout)
        self.assertIn("Configured datasets:", result.stdout)
        self.assertIn(
            "Validated report levels: portfolio, security",
            result.stdout,
        )
        self.assertIn(
            "Minimum required datasets: holdings, portfolio_performance, "
            "security_master, security_performance, transactions",
            result.stdout,
        )
        self.assertIn("Required source-data columns:", result.stdout)
        self.assertIn(
            "portfolio_performance: portfolio_id, from_date, thru_date, "
            "portfolio_return",
            result.stdout,
        )
        self.assertIn(
            "security_performance: portfolio_id, security_id, from_date, "
            "thru_date, security_return",
            result.stdout,
        )
        self.assertIn(
            "holdings: portfolio_id, security_id, holding_date, market_value",
            result.stdout,
        )
        self.assertIn(
            "transactions: portfolio_id, security_id, transaction_date, "
            "transaction_code, amount",
            result.stdout,
        )
        self.assertIn("Missing optional files: none", result.stdout)
        self.assertIn("FX rate impact methods: fx_rate", result.stdout)
        self.assertIn("Evidence-only impact methods: splits", result.stdout)
        self.assertIn("Data Issues checks enabled:", result.stdout)
        self.assertIn("duplicate_transactions", result.stdout)
        self.assertIn(
                "Data Issues policy: established checks use their configured "
                "defaults; conservative "
                "checks require explicit enablement and issue-specific scope",
                result.stdout,
            )
        self.assertIn("Transaction rules configured: 16", result.stdout)
        self.assertIn(
            "Transaction impact methods: commission, external_flow, performance, "
            "price, quantity",
            result.stdout,
        )
        self.assertIn("Transaction files checked: 2", result.stdout)
        self.assertIn("Extract contract: packaged:", result.stdout)
        self.assertIn("Enforce ambiguous Axys/APX flows: True", result.stdout)
        self.assertIn(
            "Required transaction context columns: source_destination_symbol, "
            "source_destination_type, special_security_symbol, "
            "special_security_type, transaction_security_type",
            result.stdout,
        )
        self.assertIn("Report-bundle source context:", result.stdout)
        self.assertIn("transaction semantics summary", result.stdout)
        self.assertIn(
            "Transaction codes observed: ai, by, cs, dp, dv, in, li, lo, pa, pd, "
            "rc, sa, sl, ss, ti, wd",
            result.stdout,
        )
        self.assertIn("Transaction codes without YAML rules: none", result.stdout)
        self.assertIn("Transaction semantics sources:", result.stdout)
        self.assertEqual(result.stderr, "")

    def test_validate_config_checks_available_security_view(self) -> None:
        """Security policy errors fail preflight before a standard site run."""
        with tempfile.TemporaryDirectory() as directory:
            configuration = yaml.safe_load(
                _PORTFOLIO_COMPARISON_PATH.read_text(encoding="utf-8")
            )
            packaged_directory = _PACKAGED_AXYS_APX_DATA_PATH.resolve()
            configuration["snapshots"]["a"]["path"] = str(
                packaged_directory / "snapshot_a"
            )
            configuration["snapshots"]["b"]["path"] = str(
                packaged_directory / "snapshot_b"
            )
            del configuration["security_return_impact_methods"]
            comparison_path = Path(directory) / "comparison.yaml"
            comparison_path.write_text(
                yaml.safe_dump(configuration),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(
                PpaError,
                "security_return_impact_methods is required",
            ):
                validate_config(
                    comparison_path,
                    require_complete_yaml_setup=False,
                )

    def test_validate_config_cli_rejects_incomplete_yaml_bypass(self) -> None:
        """Normal config validation does not expose an incomplete-setup bypass."""
        result = subprocess.run(
            _module_command(
                _VALIDATE_CONFIG_MODULE,
                str(_RESTATEMENT_COMPARISON_PATH),
                "--allow-incomplete-yaml",
            ),
            check=False,
            capture_output=True,
            text=True,
        )

        self.assertEqual(result.returncode, 2)
        self.assertIn("unrecognized arguments: --allow-incomplete-yaml", result.stderr)

    def test_validate_config_cli_module_reports_invalid_yaml_contract(self) -> None:
        """The CLI config validator exits nonzero for malformed YAML contracts."""
        with tempfile.TemporaryDirectory() as directory:
            configuration = _absolute_restatement_configuration()
            transaction_methods = configuration["transaction_impact_methods"]
            assert isinstance(transaction_methods, dict)
            transaction_methods["performance"] = {
                "method": "unsupported",
                "denominator_source": "begin_market_value",
            }
            comparison_path = Path(directory) / "comparison.yaml"
            comparison_path.write_text(
                yaml.safe_dump(configuration),
                encoding="utf-8",
            )

            result = subprocess.run(
                _module_command(
                    _VALIDATE_CONFIG_MODULE,
                    str(comparison_path),
                ),
                check=False,
                capture_output=True,
                text=True,
            )

        self.assertEqual(result.returncode, 1)
        self.assertEqual(result.stdout, "")
        self.assertIn("Config validation failed:", result.stderr)
        self.assertIn("performance.method must be", result.stderr)

    def test_validate_demo_matrix_script_accepts_packaged_demos(self) -> None:
        """The maintainer script confirms packaged scenario coverage."""
        result = subprocess.run(
            [sys.executable, str(_VALIDATE_DEMO_MATRIX_SCRIPT)],
            check=True,
            capture_output=True,
            text=True,
        )

        self.assertIn("Demo matrix validation passed:", result.stdout)
        self.assertIn("Demo matrix coverage includes ambiguous-flow", result.stdout)
        self.assertIn("Clean/no issue", result.stdout)
        self.assertIn("Single-restatement transaction rows", result.stdout)
        self.assertIn("Transaction rules amount explanation", result.stdout)
        self.assertIn("Context-only evidence", result.stdout)
        self.assertIn("Portfolio field-role specifications", result.stdout)
        self.assertIn("Security field-role specifications", result.stdout)
        self.assertIn("Suppressed finding", result.stdout)
        self.assertIn("Residual withheld", result.stdout)
        self.assertIn("Ambiguous flow context variants", result.stdout)
        self.assertIn("Code-only failure guard", result.stdout)
        self.assertIn("Reviewed local opt-out", result.stdout)
        self.assertIn("Review-only action quarantine", result.stdout)
        self.assertIn("Capital-return and short-side candidate gates", result.stdout)
        self.assertEqual(result.stderr, "")

    def _write_bundle(self, output_directory: Path) -> None:
        """Write a standard report bundle for CLI validation tests."""
        findings = compare_snapshots(_RESTATEMENT_COMPARISON_PATH)
        write_audit_report_bundle(
            findings,
            output_directory,
            comparison_path=_RESTATEMENT_COMPARISON_PATH,
            expand_all_supporting_files=True,
        )


def _absolute_restatement_configuration() -> dict[str, object]:
    """Return restatement YAML values with absolute fixture paths."""
    configuration = yaml.safe_load(_RESTATEMENT_COMPARISON_PATH.read_text(encoding="utf-8"))
    fixture_directory = _AXYS_SNAPSHOT_PATH.resolve()
    configuration["snapshots"]["a"]["path"] = str(fixture_directory / "axys_a")
    configuration["snapshots"]["b"]["path"] = str(
        fixture_directory / "axys_b_restatement"
    )
    packaged_configuration = yaml.safe_load(
        _PORTFOLIO_COMPARISON_PATH.read_text(encoding="utf-8")
    )
    for file_name, packaged_definition in packaged_configuration["files"].items():
        if file_name not in configuration["files"]:
            continue
        current_definition = configuration["files"][file_name]
        current_path = (
            current_definition
            if isinstance(current_definition, str)
            else current_definition["path"]
        )
        configuration["files"][file_name] = {
            "path": current_path,
            "columns": packaged_definition["columns"],
        }
    configuration["snapshots"]["a"].pop("schema", None)
    configuration["snapshots"]["b"].pop("schema", None)
    return configuration


def _write_audit_run_settings(path: Path) -> None:
    """Write the strict normal-run section used by isolated CLI tests."""
    path.write_text(
        yaml.safe_dump(
            {
                "audit": {
                    "output_directory": "output",
                    "title": None,
                    "xlsx_output": True,
                    "html_output": True,
                    "exclude_suppressed": False,
                    "reconstruction_diagnostics": False,
                    "expand_all_supporting_files": False,
                    "require_causal_attribution": False,
                }
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )


def _copy_site_snapshots(directory: Path) -> Path:
    """Copy packaged demo snapshots into a setup-style site folder."""
    site_directory = directory / "my_site_extracts"
    shutil.copytree(
        _PACKAGED_AXYS_APX_DATA_PATH / "snapshot_a",
        site_directory / "snapshot_a",
    )
    shutil.copytree(
        _PACKAGED_AXYS_APX_DATA_PATH / "snapshot_b",
        site_directory / "snapshot_b",
    )
    return site_directory


def _module_command(module_name: str, *args: str) -> list[str]:
    """Return a subprocess command that runs a package CLI module."""
    return [sys.executable, "-m", module_name, *args]


if __name__ == "__main__":
    unittest.main()
