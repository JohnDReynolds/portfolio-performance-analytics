"""Generate temporary candidate data files for the analytics demo.

This refresh helper writes only inside the temporary analytics-data-generation
workspace by default.

The generated data is meant to be believable demo data, not research-grade index
history. It uses current constituents historically, current GICS sectors, and a
simple hindsight alpha tilt so the demo can tell a clear attribution story.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
import datetime as dt
import hashlib
from io import StringIO
import json
from pathlib import Path
from typing import Final
from typing import TypeAlias
from urllib.error import URLError
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd
from openpyxl import load_workbook

try:
    import yfinance as yf
except ImportError as error:  # pragma: no cover - local generation convenience.
    raise SystemExit(
        "yfinance is required to run this refresh helper in the local .venv."
    ) from error


_WORKSPACE: Final = Path("_demo_output") / "analytics_data_generation"
_DEFAULT_OUTPUT_DIRECTORY: Final = _WORKSPACE / "generated_files"
_DEFAULT_CACHE_DIRECTORY: Final = _WORKSPACE / "cache"
_WIKIPEDIA_SP100_URL: Final = "https://en.wikipedia.org/wiki/S%26P_100"
_WIKIPEDIA_SP500_URL: Final = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
_MONTHS_PER_YEAR: Final = 12
_MIN_MONTHS: Final = 48
_DEFAULT_YEARS: Final = 10
_DEFAULT_TOP_HOLDINGS: Final = 200

_ETF_HOLDINGS_URLS: Final = {
    "spy": (
        "https://www.ssga.com/us/en/intermediary/etfs/library-content/"
        "products/fund-data/etfs/us/holdings-daily-us-en-spy.xlsx"
    ),
    "ivv": (
        "https://www.ishares.com/us/products/239726/ishares-core-sp-500-etf/"
        "1467271812596.ajax?fileType=csv&fileName=IVV_holdings&dataType=fund"
    ),
}

_SECTOR_CODES: Final = {
    "Communication Services": "CO",
    "Consumer Discretionary": "CD",
    "Consumer Staples": "CS",
    "Energy": "EN",
    "Financials": "FI",
    "Health Care": "HC",
    "Industrials": "IN",
    "Information Technology": "IT",
    "Materials": "MA",
    "Real Estate": "RE",
    "Utilities": "UT",
}

_SECTOR_ALIASES: Final = {
    "Communication": "Communication Services",
    "Communications": "Communication Services",
    "Consumer, Cyclical": "Consumer Discretionary",
    "Consumer, Non-cyclical": "Consumer Staples",
    "Financial": "Financials",
    "Healthcare": "Health Care",
    "Technology": "Information Technology",
}

Weights: TypeAlias = pd.Series | pd.DataFrame


@dataclass(frozen=True)
class Holding:
    """Current large-cap holding metadata used to build demo data."""

    ticker: str
    name: str
    sector: str
    weight: float | None = None
    shares: float | None = None


_FALLBACK_UNIVERSE: Final = (
    Holding("AAPL", "Apple Inc.", "Information Technology"),
    Holding("MSFT", "Microsoft", "Information Technology"),
    Holding("NVDA", "NVIDIA", "Information Technology"),
    Holding("AVGO", "Broadcom", "Information Technology"),
    Holding("ORCL", "Oracle", "Information Technology"),
    Holding("CRM", "Salesforce", "Information Technology"),
    Holding("ADBE", "Adobe", "Information Technology"),
    Holding("AMD", "Advanced Micro Devices", "Information Technology"),
    Holding("ACN", "Accenture", "Information Technology"),
    Holding("CSCO", "Cisco Systems", "Information Technology"),
    Holding("GOOGL", "Alphabet Class A", "Communication Services"),
    Holding("GOOG", "Alphabet Class C", "Communication Services"),
    Holding("META", "Meta Platforms", "Communication Services"),
    Holding("NFLX", "Netflix", "Communication Services"),
    Holding("DIS", "Walt Disney", "Communication Services"),
    Holding("TMUS", "T-Mobile US", "Communication Services"),
    Holding("AMZN", "Amazon.com", "Consumer Discretionary"),
    Holding("TSLA", "Tesla", "Consumer Discretionary"),
    Holding("HD", "Home Depot", "Consumer Discretionary"),
    Holding("MCD", "McDonald's", "Consumer Discretionary"),
    Holding("LOW", "Lowe's", "Consumer Discretionary"),
    Holding("NKE", "Nike", "Consumer Discretionary"),
    Holding("WMT", "Walmart", "Consumer Staples"),
    Holding("COST", "Costco", "Consumer Staples"),
    Holding("PG", "Procter & Gamble", "Consumer Staples"),
    Holding("KO", "Coca-Cola", "Consumer Staples"),
    Holding("PEP", "PepsiCo", "Consumer Staples"),
    Holding("PM", "Philip Morris International", "Consumer Staples"),
    Holding("XOM", "Exxon Mobil", "Energy"),
    Holding("CVX", "Chevron", "Energy"),
    Holding("COP", "ConocoPhillips", "Energy"),
    Holding("JPM", "JPMorgan Chase", "Financials"),
    Holding("V", "Visa", "Financials"),
    Holding("MA", "Mastercard", "Financials"),
    Holding("BAC", "Bank of America", "Financials"),
    Holding("WFC", "Wells Fargo", "Financials"),
    Holding("GS", "Goldman Sachs", "Financials"),
    Holding("BRK-B", "Berkshire Hathaway", "Financials"),
    Holding("LLY", "Eli Lilly", "Health Care"),
    Holding("UNH", "UnitedHealth Group", "Health Care"),
    Holding("JNJ", "Johnson & Johnson", "Health Care"),
    Holding("ABBV", "AbbVie", "Health Care"),
    Holding("MRK", "Merck", "Health Care"),
    Holding("TMO", "Thermo Fisher Scientific", "Health Care"),
    Holding("ABT", "Abbott Laboratories", "Health Care"),
    Holding("DHR", "Danaher", "Health Care"),
    Holding("GE", "GE Aerospace", "Industrials"),
    Holding("CAT", "Caterpillar", "Industrials"),
    Holding("RTX", "RTX", "Industrials"),
    Holding("HON", "Honeywell", "Industrials"),
    Holding("UNP", "Union Pacific", "Industrials"),
    Holding("UPS", "United Parcel Service", "Industrials"),
    Holding("BA", "Boeing", "Industrials"),
    Holding("LIN", "Linde", "Materials"),
    Holding("SHW", "Sherwin-Williams", "Materials"),
    Holding("APD", "Air Products and Chemicals", "Materials"),
    Holding("PLD", "Prologis", "Real Estate"),
    Holding("AMT", "American Tower", "Real Estate"),
    Holding("EQIX", "Equinix", "Real Estate"),
    Holding("NEE", "NextEra Energy", "Utilities"),
    Holding("SO", "Southern Company", "Utilities"),
    Holding("DUK", "Duke Energy", "Utilities"),
)


def main() -> None:
    """Run temporary demo-data generation."""
    args = _parse_args()
    args.output_directory.mkdir(parents=True, exist_ok=True)
    args.cache_directory.mkdir(parents=True, exist_ok=True)

    holdings = _load_holdings(
        cache_directory=args.cache_directory,
        refresh=args.refresh,
        holdings_source=args.holdings_source,
        holdings_url=args.holdings_url,
        seed_holdings_path=args.seed_holdings_path,
        top_holdings=args.top_holdings,
    )
    benchmark_weights = _benchmark_weights(
        holdings,
        args.cache_directory,
        args.refresh,
        args.allow_market_cap_fetch,
        args.top_holdings,
    )
    holdings = _holdings_with_weights(holdings, benchmark_weights)
    holdings = _top_holdings(holdings, args.top_holdings)
    prices = _load_monthly_prices(
        holdings,
        args.cache_directory,
        args.refresh,
        args.years,
    )
    prices = _filter_prices(prices)

    holdings = [holding for holding in holdings if holding.ticker in prices.columns]
    if len(holdings) < 25:
        raise SystemExit(f"Only {len(holdings)} usable securities found; stopping.")

    returns = prices.pct_change().dropna(how="all")
    returns = returns.dropna(axis="columns", how="any")
    holdings = [holding for holding in holdings if holding.ticker in returns.columns]
    returns = returns[[holding.ticker for holding in holdings]]

    benchmark_weights = _benchmark_weight_model(
        returns=returns,
        prices=prices,
        holdings=holdings,
        static_weights=benchmark_weights,
        cache_directory=args.cache_directory,
        refresh=args.refresh,
        years=args.years,
        model=args.benchmark_weight_model,
    )
    portfolio_weights, selected_tilt = _portfolio_weights(
        returns,
        benchmark_weights,
        holdings,
        args.alpha_tilt_multiplier,
    )

    output_paths = _write_generated_files(
        holdings=holdings,
        returns=returns,
        benchmark_weights=benchmark_weights,
        portfolio_weights=portfolio_weights,
        output_directory=args.output_directory,
    )
    summary = _summary(
        returns=returns,
        benchmark_weights=benchmark_weights,
        portfolio_weights=portfolio_weights,
        holdings=holdings,
        selected_tilt=selected_tilt,
        output_paths=output_paths,
    )
    _write_summary(summary, args.output_directory)
    _print_summary(summary)


def _parse_args() -> argparse.Namespace:
    """Parse command-line options."""
    parser = argparse.ArgumentParser(
        description="Generate temporary candidate analytics demo data.",
    )
    parser.add_argument(
        "--output-directory",
        type=Path,
        default=_DEFAULT_OUTPUT_DIRECTORY,
        help="Directory for generated temp CSV files.",
    )
    parser.add_argument(
        "--cache-directory",
        type=Path,
        default=_DEFAULT_CACHE_DIRECTORY,
        help="Directory for raw downloaded temp cache files.",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Refresh downloaded holdings, prices, and market caps.",
    )
    parser.add_argument(
        "--holdings-source",
        choices=("auto", "spy", "ivv", "sp500", "seed", "fallback"),
        default="auto",
        help="Preferred holdings source.",
    )
    parser.add_argument(
        "--holdings-url",
        help="Optional holdings CSV URL. Overrides built-in ETF holdings URLs.",
    )
    parser.add_argument(
        "--seed-holdings-path",
        type=Path,
        help="Optional seed holdings CSV with ticker, name, sector, and weight.",
    )
    parser.add_argument(
        "--top-holdings",
        type=int,
        default=_DEFAULT_TOP_HOLDINGS,
        help="Maximum number of holdings to keep by reported benchmark weight.",
    )
    parser.add_argument(
        "--years",
        type=int,
        default=_DEFAULT_YEARS,
        help="Years of daily prices to download before monthly resampling.",
    )
    parser.add_argument(
        "--allow-market-cap-fetch",
        action="store_true",
        help=(
            "Allow slower per-ticker market-cap calls when holdings lack "
            "source weights."
        ),
    )
    parser.add_argument(
        "--alpha-tilt-multiplier",
        type=float,
        default=2.0,
        help="Multiplier applied to the candidate alpha tilt strengths.",
    )
    parser.add_argument(
        "--benchmark-weight-model",
        choices=("dynamic_spy_shares", "static_source", "calibrated_static_spy"),
        default="dynamic_spy_shares",
        help=(
            "Benchmark weighting model. calibrated_static_spy infers static "
            "nonnegative weights that track SPY adjusted returns while staying "
            "near source SPY weights."
        ),
    )
    return parser.parse_args()


def _load_holdings(
    cache_directory: Path,
    refresh: bool,
    holdings_source: str,
    holdings_url: str | None,
    seed_holdings_path: Path | None,
    top_holdings: int,
) -> list[Holding]:
    """Load large-cap holdings, falling back to a curated seed list."""
    cache_path = cache_directory / f"{holdings_source}_holdings.csv"
    if cache_path.exists() and not refresh:
        return _top_holdings(_holdings_from_frame(pd.read_csv(cache_path)), top_holdings)

    if seed_holdings_path is not None:
        holdings = _holdings_from_frame(pd.read_csv(seed_holdings_path))
        _holdings_to_frame(holdings).to_csv(cache_path, index=False)
        return _top_holdings(holdings, top_holdings)

    if holdings_source in ("auto", "spy"):
        url = holdings_url or _ETF_HOLDINGS_URLS["spy"]
        try:
            holdings = _download_spy_holdings(url)
            if holdings:
                _holdings_to_frame(holdings).to_csv(cache_path, index=False)
                return _top_holdings(holdings, top_holdings)
        except Exception as error:  # pragma: no cover - network variability.
            print(f"Could not download SPY holdings; trying next source. {error}")
        if holdings_source == "spy":
            raise SystemExit("Could not load SPY holdings.")

    if holdings_source in ("auto", "ivv"):
        url = holdings_url or _ETF_HOLDINGS_URLS["ivv"]
        try:
            holdings = _download_etf_holdings(url)
            if holdings:
                _holdings_to_frame(holdings).to_csv(cache_path, index=False)
                return _top_holdings(holdings, top_holdings)
        except Exception as error:  # pragma: no cover - network variability.
            print(f"Could not download ETF holdings; trying fallback. {error}")

    if holdings_source in ("auto", "sp500"):
        sp500_cache_path = cache_directory / "sp500_holdings.csv"
        if sp500_cache_path.exists() and not refresh:
            return _top_holdings(
                _holdings_from_frame(pd.read_csv(sp500_cache_path)),
                top_holdings,
            )
        try:
            holdings = _download_sp500_holdings()
            if holdings:
                _holdings_to_frame(holdings).to_csv(sp500_cache_path, index=False)
                return _top_holdings(holdings, top_holdings)
        except Exception as error:  # pragma: no cover - network variability.
            print(f"Could not download S&P 500 holdings; trying fallback. {error}")

    if holdings_source == "auto":
        wikipedia_cache_path = cache_directory / "sp100_holdings.csv"
        if wikipedia_cache_path.exists() and not refresh:
            return _top_holdings(
                _holdings_from_frame(pd.read_csv(wikipedia_cache_path)),
                top_holdings,
            )
        try:
            holdings = _download_sp100_holdings()
            if holdings:
                _holdings_to_frame(holdings).to_csv(wikipedia_cache_path, index=False)
                return _top_holdings(holdings, top_holdings)
        except Exception as error:  # pragma: no cover - network variability.
            print(f"Could not download S&P 100 holdings; using fallback list. {error}")

    if holdings_source == "seed":
        raise SystemExit("--seed-holdings-path is required with --holdings-source seed.")
    if holdings_source == "sp500":
        raise SystemExit("Could not load S&P 500 holdings.")

    print("Using curated fallback holdings list.")
    holdings = list(_FALLBACK_UNIVERSE)
    _holdings_to_frame(holdings).to_csv(cache_path, index=False)
    return _top_holdings(holdings, top_holdings)


def _download_etf_holdings(url: str) -> list[Holding]:
    """Download and parse an ETF holdings CSV."""
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=30) as response:
        text = response.read().decode("utf-8-sig", errors="replace")
    frame = _parse_holdings_csv(text)
    return _holdings_from_frame(frame)


def _download_spy_holdings(url: str) -> list[Holding]:
    """Download SPY holdings and merge GICS sectors from S&P 500 constituents."""
    request = Request(url, headers={"User-Agent": "Mozilla/5.0", "Accept": "*/*"})
    with urlopen(request, timeout=30) as response:
        workbook_bytes = response.read()
    workbook_path = _DEFAULT_CACHE_DIRECTORY / "spy_holdings_source.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_bytes(workbook_bytes)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["holdings"] if "holdings" in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = next(
        index
        for index, row in enumerate(rows)
        if row and "Ticker" in row and "Weight" in row
    )
    headers = [str(value).strip() if value is not None else "" for value in rows[header_index]]
    records = []
    for row in rows[header_index + 1 :]:
        record = dict(zip(headers, row))
        ticker = _normalize_ticker(str(record.get("Ticker", "")))
        weight = record.get("Weight")
        if ticker and isinstance(weight, int | float):
            records.append(
                {
                    "ticker": ticker,
                    "name": str(record.get("Name", "")).strip().title(),
                    "weight": float(weight) / 100.0,
                    "shares": float(record.get("Shares Held") or np.nan),
                }
            )

    if not records:
        return []

    spy_frame = pd.DataFrame(records)
    sector_frame = _holdings_to_frame(_download_sp500_holdings())[
        ["ticker", "name", "sector"]
    ]
    merged = spy_frame.merge(
        sector_frame.rename(columns={"name": "sp500_name"}),
        on="ticker",
        how="left",
    )
    merged["sector"] = merged["sector"].fillna("")
    merged["name"] = merged["sp500_name"].fillna(merged["name"])
    return _holdings_from_frame(merged[["ticker", "name", "sector", "weight", "shares"]])


def _parse_holdings_csv(text: str) -> pd.DataFrame:
    """Parse a holdings CSV that may contain metadata rows before its header."""
    lines = text.splitlines()
    for index, line in enumerate(lines[:80]):
        columns = [column.strip().strip('"').lower() for column in line.split(",")]
        if {"ticker", "name"}.issubset(set(columns)) and (
            "sector" in columns or "gics sector" in columns
        ):
            frame = pd.read_csv(StringIO("\n".join(lines[index:])))
            return _standardize_holdings_frame(frame)
    return _standardize_holdings_frame(pd.read_csv(StringIO(text)))


def _download_sp100_holdings() -> list[Holding]:
    """Download current S&P 100 constituents from Wikipedia."""
    tables = pd.read_html(_WIKIPEDIA_SP100_URL)
    for table in tables:
        normalized = {str(column).strip().lower(): column for column in table.columns}
        ticker_column = normalized.get("symbol") or normalized.get("ticker")
        name_column = normalized.get("security") or normalized.get("name")
        sector_column = normalized.get("gics sector") or normalized.get("sector")
        if ticker_column and name_column and sector_column:
            frame = table[[ticker_column, name_column, sector_column]].rename(
                columns={
                    ticker_column: "ticker",
                    name_column: "name",
                    sector_column: "sector",
                }
            )
            return _holdings_from_frame(frame)
    return []


def _download_sp500_holdings() -> list[Holding]:
    """Download current S&P 500 constituents from Wikipedia."""
    request = Request(_WIKIPEDIA_SP500_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=30) as response:
        text = response.read().decode("utf-8", errors="replace")
    tables = pd.read_html(StringIO(text))
    for table in tables:
        normalized = {str(column).strip().lower(): column for column in table.columns}
        ticker_column = normalized.get("symbol") or normalized.get("ticker")
        name_column = normalized.get("security") or normalized.get("name")
        sector_column = normalized.get("gics sector") or normalized.get("sector")
        if ticker_column and name_column and sector_column:
            frame = table[[ticker_column, name_column, sector_column]].rename(
                columns={
                    ticker_column: "ticker",
                    name_column: "name",
                    sector_column: "sector",
                }
            )
            return _holdings_from_frame(frame)
    return []


def _holdings_from_frame(frame: pd.DataFrame) -> list[Holding]:
    """Convert a holdings DataFrame into validated holdings."""
    frame = _standardize_holdings_frame(frame)
    holdings: list[Holding] = []
    for row in frame.itertuples(index=False):
        ticker = _normalize_ticker(str(row.ticker))
        sector = _normalize_sector(str(row.sector))
        weight = getattr(row, "weight", None)
        weight = None if pd.isna(weight) else float(weight)
        shares = getattr(row, "shares", None)
        shares = None if pd.isna(shares) else float(shares)
        if ticker and sector in _SECTOR_CODES:
            holdings.append(Holding(ticker, str(row.name).strip(), sector, weight, shares))
    return holdings


def _standardize_holdings_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Return holdings columns as ticker, name, sector, and optional weight."""
    normalized = {_normalize_column_name(column): column for column in frame.columns}
    ticker_column = normalized.get("ticker") or normalized.get("symbol")
    name_column = (
        normalized.get("name")
        or normalized.get("security")
        or normalized.get("company")
        or normalized.get("holding")
    )
    sector_column = normalized.get("sector") or normalized.get("gics_sector")
    weight_column = (
        normalized.get("weight")
        or normalized.get("weight_pct")
        or normalized.get("weight_percent")
        or normalized.get("weight_percentage")
    )
    if ticker_column is None or name_column is None or sector_column is None:
        raise ValueError(
            "Holdings data must contain ticker/symbol, name/security, and sector columns."
        )

    output = pd.DataFrame(
        {
            "ticker": frame[ticker_column].map(_normalize_ticker),
            "name": frame[name_column].astype(str).str.strip(),
            "sector": frame[sector_column].map(lambda value: _normalize_sector(str(value))),
        }
    )
    if weight_column is not None:
        output["weight"] = frame[weight_column].map(_parse_weight)
    else:
        output["weight"] = np.nan
    shares_column = normalized.get("shares") or normalized.get("shares_held")
    if shares_column is not None:
        output["shares"] = pd.to_numeric(frame[shares_column], errors="coerce")
    else:
        output["shares"] = np.nan

    output = output.dropna(subset=["ticker", "name", "sector"])
    output = output[output["ticker"] != ""]
    output = output[output["sector"].isin(_SECTOR_CODES)]
    return output.drop_duplicates("ticker")


def _normalize_column_name(column: object) -> str:
    """Normalize source column names for holdings parsing."""
    return (
        str(column)
        .strip()
        .lower()
        .replace("%", "pct")
        .replace("(", "")
        .replace(")", "")
        .replace("/", "_")
        .replace("-", "_")
        .replace(" ", "_")
    )


def _parse_weight(value: object) -> float:
    """Parse a holdings weight as a decimal or percentage."""
    if value is None or pd.isna(value):
        return np.nan
    text = str(value).replace("%", "").replace(",", "").strip()
    if not text or text == "-":
        return np.nan
    number = float(text)
    return number / 100.0 if number > 1.0 else number


def _normalize_sector(sector: str) -> str:
    """Normalize sector labels to the GICS names used by the demo."""
    sector = sector.strip()
    return _SECTOR_ALIASES.get(sector, sector)


def _holdings_to_frame(holdings: list[Holding]) -> pd.DataFrame:
    """Return holdings as a cache-friendly DataFrame."""
    return pd.DataFrame(
        [
            {
                "ticker": holding.ticker,
                "name": holding.name,
                "sector": holding.sector,
                "weight": holding.weight,
                "shares": holding.shares,
            }
            for holding in holdings
        ]
    )


def _top_holdings(holdings: list[Holding], top_holdings: int) -> list[Holding]:
    """Return top holdings by reported weight when weights are available."""
    if top_holdings <= 0 or len(holdings) <= top_holdings:
        return holdings
    weighted = [holding for holding in holdings if holding.weight is not None]
    if len(weighted) >= top_holdings:
        return sorted(weighted, key=lambda holding: holding.weight or 0.0, reverse=True)[
            :top_holdings
        ]
    return holdings


def _holdings_with_weights(
    holdings: list[Holding],
    weights: pd.Series,
) -> list[Holding]:
    """Attach benchmark weights to holdings when source weights are absent."""
    output: list[Holding] = []
    for holding in holdings:
        weight = holding.weight
        if weight is None and holding.ticker in weights.index:
            weight = float(weights[holding.ticker])
        output.append(
            Holding(holding.ticker, holding.name, holding.sector, weight, holding.shares)
        )
    return output


def _normalize_ticker(ticker: str) -> str:
    """Normalize tickers for Yahoo Finance."""
    return ticker.strip().upper().replace(".", "-")


def _load_monthly_prices(
    holdings: list[Holding],
    cache_directory: Path,
    refresh: bool,
    years: int,
) -> pd.DataFrame:
    """Download or load cached monthly adjusted close prices."""
    tickers = [holding.ticker for holding in holdings]
    ticker_hash = hashlib.sha256(",".join(tickers).encode("utf-8")).hexdigest()[:12]
    cache_path = cache_directory / f"monthly_prices_{years}y_{ticker_hash}.csv"
    if cache_path.exists() and not refresh:
        return pd.read_csv(cache_path, index_col=0, parse_dates=True)

    raw = yf.download(
        tickers=tickers,
        period=f"{years}y",
        interval="1d",
        auto_adjust=True,
        progress=False,
        threads=True,
    )
    if isinstance(raw.columns, pd.MultiIndex):
        if "Close" in raw.columns.get_level_values(0):
            close = raw["Close"]
        else:
            close = raw.xs("Close", axis=1, level=-1)
    else:
        close = raw[["Close"]].rename(columns={"Close": tickers[0]})

    monthly = close.resample("ME").last().dropna(how="all")
    monthly.to_csv(cache_path)
    return monthly


def _filter_prices(prices: pd.DataFrame) -> pd.DataFrame:
    """Keep securities with enough monthly price history."""
    prices = prices.loc[prices.index <= _last_complete_month_end()]
    prices = prices.dropna(axis="columns", thresh=_MIN_MONTHS)
    prices = prices.dropna(axis="rows", how="all")
    return prices


def _last_complete_month_end() -> pd.Timestamp:
    """Return the most recent completed calendar month-end date."""
    first_day_of_current_month = pd.Timestamp.today().normalize().replace(day=1)
    return first_day_of_current_month - pd.Timedelta(days=1)


def _benchmark_weights(
    holdings: list[Holding],
    cache_directory: Path,
    refresh: bool,
    allow_market_cap_fetch: bool,
    top_holdings: int,
) -> pd.Series:
    """Return market-cap-style benchmark weights with an equal-weight fallback."""
    provided_weights = pd.Series(
        {
            holding.ticker: holding.weight
            for holding in holdings
            if holding.weight is not None
        },
        dtype="float64",
    )
    if provided_weights.notna().sum() >= max(10, len(holdings) // 2):
        print("Using source-provided cap weights from holdings data.")
        return provided_weights

    if not allow_market_cap_fetch and len(holdings) > top_holdings:
        print(
            "Holdings do not include source weights. Using equal weights for this "
            "temporary run; provide --seed-holdings-path for a real weighted "
            "benchmark or use --allow-market-cap-fetch for slower market-cap calls."
        )
        return pd.Series(1.0, index=[holding.ticker for holding in holdings])

    cache_path = cache_directory / "market_caps.csv"
    if cache_path.exists() and not refresh:
        frame = pd.read_csv(cache_path)
        print("Using cached market-cap weights.")
        return pd.Series(frame["market_cap"].to_numpy(), index=frame["ticker"])

    rows: list[dict[str, float | str]] = []
    for holding in holdings:
        market_cap = np.nan
        try:
            ticker = yf.Ticker(holding.ticker)
            fast_info = ticker.fast_info
            market_cap = float(
                fast_info.get("marketCap")
                or fast_info.get("market_cap")
                or np.nan
            )
        except Exception:  # pragma: no cover - network variability.
            market_cap = np.nan
        rows.append({"ticker": holding.ticker, "market_cap": market_cap})

    frame = pd.DataFrame(rows)
    if frame["market_cap"].notna().sum() < max(10, len(holdings) // 2):
        print("Market-cap coverage was insufficient; using equal weights.")
        frame["market_cap"] = 1.0
    else:
        print("Using current market-cap weights from yfinance.")
    frame.to_csv(cache_path, index=False)
    return pd.Series(frame["market_cap"].to_numpy(), index=frame["ticker"])


def _normalize_weights(weights: pd.Series) -> pd.Series:
    """Normalize nonnegative weights to sum to 1.0."""
    weights = weights.clip(lower=0.0).fillna(0.0)
    total = weights.sum()
    if total <= 0:
        return pd.Series(1.0 / len(weights), index=weights.index)
    return weights / total


def _benchmark_weight_model(
    returns: pd.DataFrame,
    prices: pd.DataFrame,
    holdings: list[Holding],
    static_weights: pd.Series,
    cache_directory: Path,
    refresh: bool,
    years: int,
    model: str,
) -> Weights:
    """Return benchmark weights using the requested temporary benchmark model."""
    static_weights = static_weights.reindex(returns.columns).fillna(0.0)
    static_weights = _normalize_weights(static_weights)
    if model == "static_source":
        print("Using static source weights.")
        return static_weights

    if model == "calibrated_static_spy":
        target = _load_spy_target_returns(cache_directory, refresh, years, returns.index)
        weights = _calibrate_static_spy_weights(
            returns=returns,
            target=target,
            prior_weights=static_weights,
        )
        print(
            "Using calibrated static benchmark weights fitted to SPY adjusted "
            "returns. This is a proxy for historical beginning weights."
        )
        return weights

    shares = pd.Series(
        {
            holding.ticker: holding.shares
            for holding in holdings
            if holding.shares is not None
        },
        dtype="float64",
    ).reindex(returns.columns)
    if shares.notna().sum() >= max(25, len(returns.columns) // 2):
        begin_values = prices[returns.columns].shift(1).loc[returns.index].mul(
            shares,
            axis="columns",
        )
        weights = begin_values.div(begin_values.sum(axis="columns"), axis="rows")
        print("Using dynamic beginning-of-period cap weights from SPY shares.")
        return weights.fillna(0.0)

    print("Using static source weights.")
    return static_weights


def _load_spy_target_returns(
    cache_directory: Path,
    refresh: bool,
    years: int,
    index: pd.DatetimeIndex,
) -> pd.Series:
    """Load SPY adjusted-close monthly returns for benchmark calibration."""
    cache_path = cache_directory / f"spy_target_returns_{years}y.csv"
    if cache_path.exists() and not refresh and cache_path.stat().st_size > 20:
        target = pd.read_csv(cache_path, index_col=0, parse_dates=True).iloc[:, 0]
    else:
        raw = yf.download(
            tickers=["SPY"],
            period=f"{years}y",
            interval="1d",
            auto_adjust=True,
            progress=False,
            threads=False,
        )
        close = raw["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
        target = close.resample("ME").last().pct_change().dropna()
        if target.empty:
            raise SystemExit("SPY target return download did not return usable data.")
        target.to_frame("return").to_csv(cache_path)

    aligned = target.reindex(index).dropna()
    if aligned.empty:
        raise SystemExit("SPY target returns do not overlap generated return periods.")
    return aligned


def _calibrate_static_spy_weights(
    returns: pd.DataFrame,
    target: pd.Series,
    prior_weights: pd.Series,
) -> pd.Series:
    """Infer static nonnegative weights that match SPY-level returns.

    The calibration keeps the temporary benchmark cap-weight-like, but it avoids
    the extreme hindsight bias from applying today's mega-cap weights across the
    entire lookback window.
    """
    from scipy.optimize import minimize

    common_index = returns.index.intersection(target.index)
    aligned_returns = returns.loc[common_index]
    aligned_target = target.loc[common_index]
    prior_weights = _normalize_weights(
        prior_weights.reindex(aligned_returns.columns).fillna(0.0)
    )
    matrix = aligned_returns.to_numpy()
    target_values = aligned_target.to_numpy()
    prior_values = prior_weights.to_numpy()
    target_cumulative = np.prod(1.0 + target_values) - 1.0
    cumulative_strength = 5.0
    prior_strength = 0.001

    def objective(weights: np.ndarray) -> float:
        candidate = matrix.dot(weights)
        tracking_error = np.mean((candidate - target_values) ** 2)
        prior_distance = np.mean((weights - prior_values) ** 2)
        candidate_cumulative = np.prod(1.0 + candidate) - 1.0
        cumulative_error = (candidate_cumulative - target_cumulative) ** 2
        return (
            tracking_error
            + prior_strength * prior_distance
            + cumulative_strength * cumulative_error
        )

    starts = [
        prior_values,
        np.repeat(1.0 / len(prior_values), len(prior_values)),
        np.sqrt(prior_values) / np.sqrt(prior_values).sum(),
    ]
    constraints = [{"type": "eq", "fun": lambda weights: weights.sum() - 1.0}]
    bounds = [(0.0, 1.0) for _ in prior_values]
    best_result = None
    for start in starts:
        result = minimize(
            objective,
            start,
            method="SLSQP",
            bounds=bounds,
            constraints=constraints,
            options={"maxiter": 2000, "ftol": 1e-12},
        )
        if best_result is None or result.fun < best_result.fun:
            best_result = result

    if best_result is None or not best_result.success:
        message = "unknown optimizer failure" if best_result is None else best_result.message
        raise SystemExit(f"Could not calibrate SPY benchmark weights: {message}")

    return pd.Series(best_result.x, index=aligned_returns.columns)


def _portfolio_weights(
    returns: pd.DataFrame,
    benchmark_weights: Weights,
    holdings: list[Holding],
    alpha_tilt_multiplier: float,
) -> tuple[Weights, float]:
    """Create an alpha-tilted portfolio weight vector within GICS sectors."""
    raw_score = returns.mean() / returns.std(ddof=0).replace(0, np.nan)
    raw_score = raw_score.replace([np.inf, -np.inf], np.nan).fillna(0.0)
    sector_by_ticker = {holding.ticker: holding.sector for holding in holdings}
    score = pd.Series(0.0, index=returns.columns)
    for sector in sorted({sector_by_ticker[ticker] for ticker in returns.columns}):
        tickers = [
            ticker
            for ticker in returns.columns
            if sector_by_ticker.get(ticker) == sector
        ]
        sector_score = raw_score.loc[tickers]
        denominator = sector_score.std(ddof=0)
        if denominator == 0.0 or np.isnan(denominator):
            score.loc[tickers] = 0.0
        else:
            score.loc[tickers] = (sector_score - sector_score.mean()) / denominator

    best_weights = benchmark_weights.copy()
    best_tilt = 0.0
    benchmark_return = _period_returns(returns, benchmark_weights)
    benchmark_sharpe = _annualized_sharpe(benchmark_return)
    best_active_return = -np.inf

    for base_tilt in np.linspace(0.10, 1.50, 15):
        tilt = base_tilt * alpha_tilt_multiplier
        if isinstance(benchmark_weights, pd.DataFrame):
            candidate = benchmark_weights.mul(1.0 + tilt * score, axis="columns")
        else:
            candidate = benchmark_weights * (1.0 + tilt * score)
        candidate = _normalize_within_benchmark_sectors(
            candidate,
            benchmark_weights,
            sector_by_ticker,
        )
        candidate_return = _period_returns(returns, candidate)
        active_return = _cumulative_return(candidate_return) - _cumulative_return(
            benchmark_return
        )
        sharpe = _annualized_sharpe(candidate_return)
        if active_return > best_active_return:
            best_weights = candidate
            best_tilt = float(tilt)
            best_active_return = active_return
        if active_return > 0 and sharpe > benchmark_sharpe:
            return candidate, float(tilt)

    return best_weights, best_tilt


def _normalize_within_benchmark_sectors(
    candidate: Weights,
    benchmark_weights: Weights,
    sector_by_ticker: dict[str, str],
) -> Weights:
    """Normalize weights while preserving benchmark sector weights."""
    if isinstance(candidate, pd.DataFrame):
        return _normalize_weight_frame_within_benchmark_sectors(
            candidate,
            benchmark_weights,
            sector_by_ticker,
        )

    candidate = candidate.clip(lower=0.0).fillna(0.0)
    assert isinstance(benchmark_weights, pd.Series)
    output = pd.Series(0.0, index=candidate.index)
    for sector in sorted({sector_by_ticker[ticker] for ticker in candidate.index}):
        tickers = [
            ticker
            for ticker in candidate.index
            if sector_by_ticker.get(ticker) == sector
        ]
        sector_total = benchmark_weights.loc[tickers].sum()
        candidate_total = candidate.loc[tickers].sum()
        if candidate_total <= 0.0:
            output.loc[tickers] = benchmark_weights.loc[tickers]
        else:
            output.loc[tickers] = candidate.loc[tickers] / candidate_total * sector_total
    return _normalize_weights(output)


def _normalize_weight_frame_within_benchmark_sectors(
    candidate: pd.DataFrame,
    benchmark_weights: Weights,
    sector_by_ticker: dict[str, str],
) -> pd.DataFrame:
    """Normalize each period while preserving benchmark sector weights."""
    candidate = candidate.clip(lower=0.0).fillna(0.0)
    assert isinstance(benchmark_weights, pd.DataFrame)
    output = pd.DataFrame(0.0, index=candidate.index, columns=candidate.columns)
    for sector in sorted({sector_by_ticker[ticker] for ticker in candidate.columns}):
        tickers = [
            ticker
            for ticker in candidate.columns
            if sector_by_ticker.get(ticker) == sector
        ]
        sector_total = benchmark_weights[tickers].sum(axis="columns")
        candidate_total = candidate[tickers].sum(axis="columns")
        fallback = benchmark_weights[tickers]
        scaled = candidate[tickers].div(candidate_total, axis="rows").mul(
            sector_total,
            axis="rows",
        )
        output[tickers] = scaled.where(candidate_total > 0.0, fallback)
    return output.div(output.sum(axis="columns"), axis="rows").fillna(0.0)


def _write_generated_files(
    holdings: list[Holding],
    returns: pd.DataFrame,
    benchmark_weights: pd.Series,
    portfolio_weights: pd.Series,
    output_directory: Path,
) -> dict[str, str]:
    """Write generated performance, classification, and mapping CSV files."""
    performance_directory = output_directory / "performance"
    classification_directory = output_directory / "classifications"
    mapping_directory = output_directory / "mappings"
    for directory in (performance_directory, classification_directory, mapping_directory):
        directory.mkdir(parents=True, exist_ok=True)

    holding_by_ticker = {holding.ticker: holding for holding in holdings}
    benchmark_path = performance_directory / "Generated Large-Cap Benchmark.csv"
    portfolio_path = performance_directory / "Generated Large-Cap Alpha Portfolio.csv"
    security_path = classification_directory / "Generated Security.csv"
    sector_path = classification_directory / "Generated Economic Sector.csv"
    mapping_path = (
        mapping_directory
        / "Generated Security--to--Generated Economic Sector.csv"
    )

    _write_performance(benchmark_path, returns, benchmark_weights, holding_by_ticker)
    _write_performance(portfolio_path, returns, portfolio_weights, holding_by_ticker)
    _write_security_classification(security_path, holdings)
    _write_sector_classification(sector_path, holdings)
    _write_sector_mapping(mapping_path, holdings)

    return {
        "benchmark_performance": str(benchmark_path),
        "portfolio_performance": str(portfolio_path),
        "security_classification": str(security_path),
        "sector_classification": str(sector_path),
        "security_to_sector_mapping": str(mapping_path),
    }


def _write_performance(
    path: Path,
    returns: pd.DataFrame,
    weights: Weights,
    holding_by_ticker: dict[str, Holding],
) -> None:
    """Write one ppar narrow-format performance CSV."""
    rows: list[dict[str, object]] = []
    previous_date: pd.Timestamp | None = None
    for period_end, period_returns in returns.iterrows():
        if previous_date is None:
            previous_date = period_end - pd.offsets.MonthEnd(1)
        period_start = (previous_date + pd.Timedelta(days=1)).date()
        thru_date = period_end.date()
        for ticker, period_return in period_returns.items():
            if isinstance(weights, pd.DataFrame):
                weight = weights.at[period_end, ticker]
            else:
                weight = weights[ticker]
            rows.append(
                {
                    "from_date": period_start,
                    "thru_date": thru_date,
                    "identifier": ticker,
                    "weight": round(float(weight), 12),
                    "return": round(float(period_return), 12),
                    "name": holding_by_ticker[ticker].name,
                }
            )
        previous_date = period_end

    pd.DataFrame(rows).to_csv(path, index=False)


def _write_security_classification(path: Path, holdings: list[Holding]) -> None:
    """Write security identifier-to-name CSV without headers."""
    frame = pd.DataFrame(
        [(holding.ticker, holding.name) for holding in holdings],
        columns=["identifier", "name"],
    )
    frame.to_csv(path, index=False, header=False)


def _write_sector_classification(path: Path, holdings: list[Holding]) -> None:
    """Write GICS sector code-to-name CSV without headers."""
    sectors = sorted({holding.sector for holding in holdings})
    frame = pd.DataFrame(
        [(_SECTOR_CODES[sector], sector) for sector in sectors],
        columns=["identifier", "name"],
    )
    frame.to_csv(path, index=False, header=False)


def _write_sector_mapping(path: Path, holdings: list[Holding]) -> None:
    """Write security-to-GICS-sector mapping CSV without headers."""
    frame = pd.DataFrame(
        [(holding.ticker, _SECTOR_CODES[holding.sector]) for holding in holdings],
        columns=["security", "sector"],
    )
    frame.to_csv(path, index=False, header=False)


def _summary(
    returns: pd.DataFrame,
    benchmark_weights: Weights,
    portfolio_weights: Weights,
    holdings: list[Holding],
    selected_tilt: float,
    output_paths: dict[str, str],
) -> dict[str, object]:
    """Return generation summary metrics."""
    benchmark_return = _period_returns(returns, benchmark_weights)
    portfolio_return = _period_returns(returns, portfolio_weights)
    active_weight = _average_weights(portfolio_weights) - _average_weights(
        benchmark_weights
    )
    active_contribution = active_weight * returns.mean() * _MONTHS_PER_YEAR
    holding_by_ticker = {holding.ticker: holding for holding in holdings}

    top_active = active_contribution.sort_values(ascending=False).head(10)
    top_rows = [
        {
            "ticker": ticker,
            "name": holding_by_ticker[ticker].name,
            "sector": holding_by_ticker[ticker].sector,
            "annualized_active_contribution": round(float(value), 6),
        }
        for ticker, value in top_active.items()
    ]

    return {
        "generated_at": dt.datetime.now(dt.UTC).isoformat(),
        "usable_security_count": len(returns.columns),
        "period_count": len(returns),
        "from_date": str(returns.index[0].date()),
        "thru_date": str(returns.index[-1].date()),
        "selected_tilt": round(selected_tilt, 4),
        "portfolio_cumulative_return": round(
            _cumulative_return(portfolio_return), 6
        ),
        "benchmark_cumulative_return": round(
            _cumulative_return(benchmark_return), 6
        ),
        "active_return": round(
            _cumulative_return(portfolio_return)
            - _cumulative_return(benchmark_return),
            6,
        ),
        "portfolio_annualized_sharpe": round(
            _annualized_sharpe(portfolio_return), 6
        ),
        "benchmark_annualized_sharpe": round(
            _annualized_sharpe(benchmark_return), 6
        ),
        "top_active_contributors": top_rows,
        "output_paths": output_paths,
    }


def _period_returns(returns: pd.DataFrame, weights: Weights) -> pd.Series:
    """Return weighted period returns."""
    if isinstance(weights, pd.DataFrame):
        return returns.mul(weights).sum(axis="columns")
    return returns.mul(weights, axis="columns").sum(axis="columns")


def _average_weights(weights: Weights) -> pd.Series:
    """Return one average weight vector for contribution summaries."""
    if isinstance(weights, pd.DataFrame):
        return weights.mean(axis=0)
    return weights


def _cumulative_return(period_returns: pd.Series) -> float:
    """Return cumulative compound return."""
    return float((1.0 + period_returns).prod() - 1.0)


def _annualized_sharpe(period_returns: pd.Series) -> float:
    """Return annualized Sharpe ratio using zero risk-free rate."""
    standard_deviation = float(period_returns.std(ddof=0))
    if standard_deviation == 0.0:
        return 0.0
    return float(period_returns.mean() / standard_deviation * np.sqrt(_MONTHS_PER_YEAR))


def _write_summary(summary: dict[str, object], output_directory: Path) -> None:
    """Write summary metrics next to generated files."""
    (output_directory / "summary.json").write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )


def _print_summary(summary: dict[str, object]) -> None:
    """Print a concise generation summary."""
    print("Generated temporary analytics demo data")
    print(f"Usable securities: {summary['usable_security_count']}")
    print(f"Periods: {summary['period_count']}")
    print(f"Date range: {summary['from_date']} to {summary['thru_date']}")
    print(f"Portfolio cumulative return: {summary['portfolio_cumulative_return']}")
    print(f"Benchmark cumulative return: {summary['benchmark_cumulative_return']}")
    print(f"Active return: {summary['active_return']}")
    print(f"Portfolio Sharpe: {summary['portfolio_annualized_sharpe']}")
    print(f"Benchmark Sharpe: {summary['benchmark_annualized_sharpe']}")
    print("Output files:")
    for path in summary["output_paths"].values():
        print(f"- {path}")


if __name__ == "__main__":
    main()
