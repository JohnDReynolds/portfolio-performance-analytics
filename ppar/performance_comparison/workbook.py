"""Excel workbook presentation for performance comparison review."""

from __future__ import annotations

# Python imports
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
import datetime as dt
from pathlib import Path
from typing import Any

# Third-party imports
import polars as pl

# Project imports
import ppar.utilities as util
from ppar.errors import PpaError
from ppar.performance_comparison import schema as pc_cols
from ppar.performance_comparison import explain as _pc_explain
from ppar.performance_comparison import findings as _pc_findings
from ppar.performance_comparison import rendering as _pc_rendering
from ppar.performance_comparison import review_model as _pc_review_model

REVIEW_WORKBOOK_ARTIFACT = _pc_review_model.REVIEW_WORKBOOK_ARTIFACT
REVIEW_WORKBOOK_FILE_NAME = _pc_review_model.REVIEW_WORKBOOK_FILE_NAME
WORKBOOK_NUMBER_FORMAT = "0.000000"
_MINIMUM_COLUMN_WIDTHS = {
    _pc_findings.SNAPSHOT_A_VALUE: 16,
    _pc_findings.SNAPSHOT_B_VALUE: 16,
    _pc_findings.DELTA_B_MINUS_A: 16,
    "change": 16,
}
_EXCEL_HEADER_LINE_BREAKS = {
    "Performance Difference": "Performance\nDifference",
    "Explained Difference": "Explained\nDifference",
    "Unexplained Difference": "Unexplained\nDifference",
    "Dataset Field": "Dataset\nField",
    "Input Dataset": "Input\nDataset",
    "Input Field": "Input\nField",
    "Snapshot A Value": "Snapshot A\nValue",
    "Snapshot B Value": "Snapshot B\nValue",
    "B - A Difference": "B - A\nDifference",
    "Performance Difference Explained": "Performance\nDifference\nExplained",
    "Related Performance Difference": "Related\nPerformance\nDifference",
    "Review Guidance": "Review\nGuidance",
    "Review Key": "Review\nKey",
    "What Changed": "What\nChanged",
}

PRIMARY_DIFFERENCE_SHEETS = _pc_review_model.PRIMARY_REVIEW_SHEETS
SHARED_REVIEW_SHEETS = _pc_review_model.SHARED_REVIEW_SHEETS
EXPECTED_SHEETS = _pc_review_model.EXPECTED_REVIEW_SHEETS
REQUIRED_HEADERS = {
    _pc_review_model.PERFORMANCE_DIFFERENCES_SHEET: (
        "Portfolio",
        "From Date",
        "Thru Date",
        "Performance Difference",
        "Explained Difference",
        "Unexplained Difference",
        "Status",
        "Comments",
        "Review Key",
    ),
    _pc_review_model.IDENTIFIABLE_CAUSES_SHEET: (
        "Portfolio",
        "From Date",
        "Thru Date",
        "As Of Date",
        "Dataset Field",
        "Security",
        "Snapshot A Value",
        "Snapshot B Value",
        "B - A Difference",
        "Performance Difference Explained",
        "Related Performance Difference",
        "Review Guidance",
        "Review Key",
    ),
    _pc_review_model.OTHER_EVIDENCE_SHEET: (
        "Portfolio",
        "From Date",
        "Thru Date",
        "Input Dataset",
        "Input Field",
        "Security",
        "Snapshot A Value",
        "Snapshot B Value",
        "B - A Difference",
        "What Changed",
        "Review Guidance",
        "Review Key",
    ),
    _pc_review_model.RAW_AUDIT_TRAIL_SHEET: (
        "Portfolio",
        "From Date",
        "Thru Date",
        "Input Dataset",
        "Input Field",
        "Security",
        "Message",
        "Review Key",
    ),
}


@dataclass(frozen=True)
class ReviewWorkbookSheet:
    """Describe one review workbook worksheet.

    Attributes:
        artifact_name: Stable bundle artifact key for the sheet.
        sheet_name: Excel worksheet name.
        table: DataFrame containing sheet rows.
        columns: Optional ordered internal column names to display.
        labels: Optional mapping from internal column names to display headers.
    """

    artifact_name: str
    sheet_name: str
    table: pl.DataFrame
    columns: tuple[str, ...] | None = None
    labels: Mapping[str, str] | None = None


def ensure_openpyxl_installed() -> None:
    """Raise a clear error if the Excel workbook dependency is unavailable.

    Raises:
        PpaError: If ``openpyxl`` is not installed.
    """
    _load_openpyxl()


def write_review_workbook_sheets(
    sheets: Sequence[ReviewWorkbookSheet],
    output_path: util.PathLike,
    *,
    column_tooltip: Callable[[str], str],
) -> Path:
    """Write workbook sheets to an XLSX file.

    Args:
        sheets: Ordered worksheet specifications.
        output_path: Destination workbook path. Parent directories are created.
        column_tooltip: Callback returning header comments for internal columns.

    Returns:
        Normalized workbook path.

    Raises:
        PpaError: If the Excel workbook dependency is not installed.
    """
    workbook_class, styles = _load_openpyxl()
    workbook = workbook_class()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet in sheets:
        _add_workbook_sheet(workbook, sheet, styles, column_tooltip)

    workbook_path = Path(output_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    return workbook_path


def workbook_artifact_issues(
    bundle_path: Path,
    artifacts: Mapping[str, object],
) -> list[str]:
    """Return optional XLSX review workbook validation issues."""
    artifact_file = artifacts.get(REVIEW_WORKBOOK_ARTIFACT)
    if artifact_file is None:
        return []
    if not isinstance(artifact_file, str) or not artifact_file:
        return ["manifest artifact 'review_workbook' is malformed"]

    workbook_path = bundle_path / artifact_file
    if not workbook_path.is_file():
        return [f"artifact file {artifact_file!r} is missing"]

    try:
        # pylint: disable=import-outside-toplevel
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        return [
            "report.xlsx cannot be validated because dependency 'openpyxl' "
            "is not installed"
        ]

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as error:  # pylint: disable=broad-exception-caught
        return [f"report.xlsx could not be opened: {error}"]

    issues = _review_workbook_sheet_issues(workbook)
    workbook.close()
    return issues


def _load_openpyxl() -> tuple[type[Any], dict[str, Any]]:
    """Return openpyxl classes or raise a clear dependency error."""
    try:
        # pylint: disable=import-outside-toplevel
        from openpyxl import Workbook
        from openpyxl.comments import Comment  # type: ignore[import-untyped]
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
    except ImportError as error:
        raise PpaError(
            "XLSX review workbook export requires dependency 'openpyxl'. "
            "Install the package with its runtime dependencies.",
            None,
        ) from error
    return (
        Workbook,
        {
            "header_font": Font(bold=True, color="FFFFFF"),
            "header_fill": PatternFill(
                fill_type="solid",
                start_color="1F4E78",
                end_color="1F4E78",
            ),
            "header_alignment": Alignment(wrap_text=True, vertical="top"),
            "comment_class": Comment,
        },
    )


def _add_workbook_sheet(
    workbook: Any,
    sheet: ReviewWorkbookSheet,
    styles: Mapping[str, Any],
    column_tooltip: Callable[[str], str],
) -> None:
    """Add one formatted worksheet to a workbook."""
    worksheet = workbook.create_sheet(sheet.sheet_name)
    table = sheet.table
    columns = _workbook_sheet_columns(sheet)
    headers = [
        _excel_header_label(_workbook_column_label(column, sheet.labels))
        for column in columns
    ]
    worksheet.append(headers)
    for row in table.select(columns).iter_rows(named=True):
        worksheet.append(
            [
                _workbook_cell_value(row[column], column_name=column)
                for column in columns
            ]
        )

    worksheet.freeze_panes = "A2"
    max_column_letter = worksheet.cell(row=1, column=len(columns)).column_letter
    worksheet.auto_filter.ref = f"A1:{max_column_letter}{max(worksheet.max_row, 1)}"
    for column_name, cell in zip(columns, worksheet[1]):
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_alignment"]
        cell.comment = styles["comment_class"](column_tooltip(column_name), "ppar")
    worksheet.row_dimensions[1].height = 36
    _format_workbook_columns(worksheet, columns, headers)


def _workbook_sheet_columns(sheet: ReviewWorkbookSheet) -> tuple[str, ...]:
    """Return columns for a workbook sheet, preserving configured order."""
    if sheet.columns is not None:
        return tuple(column for column in sheet.columns if column in sheet.table.columns)
    return tuple(sheet.table.columns)


def _workbook_column_label(column: str, labels: Mapping[str, str] | None) -> str:
    """Return the display label for a workbook column."""
    if labels is None:
        return _pc_rendering.display_header(column)
    return labels.get(column, _pc_rendering.display_header(column))


def _excel_header_label(header: str) -> str:
    """Return an Excel header that wraps only between words."""
    return _EXCEL_HEADER_LINE_BREAKS.get(header, header)


def _workbook_cell_value(value: object, *, column_name: str) -> object:
    """Return a scalar value suitable for openpyxl cells."""
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        return round(float(value), 6)
    if isinstance(value, str):
        numeric_value = _workbook_number_from_text(value)
        if numeric_value is not None:
            return numeric_value
        return value
    if isinstance(value, (dt.date, dt.datetime, int, float, str, bool)) or value is None:
        return value
    return _format_value(value)


def _workbook_number_from_text(value: str) -> float | None:
    """Return a rounded workbook number for plain numeric text."""
    stripped_value = value.strip()
    if stripped_value.startswith("'"):
        stripped_value = stripped_value[1:].strip()
    if not stripped_value:
        return None
    try:
        numeric_value = float(stripped_value)
    except ValueError:
        return None
    if numeric_value in {float("inf"), float("-inf")} or numeric_value != numeric_value:
        return None
    return round(numeric_value, 6)


def _format_workbook_columns(
    worksheet: Any,
    columns: Sequence[str],
    headers: Sequence[str],
) -> None:
    """Apply readable widths and common number formats to a worksheet."""
    for column_index, (column_name, header) in enumerate(zip(columns, headers), start=1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        max_width = 0
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            max_width = max(max_width, len(_format_value(cell.value)))
            if column_name in {
                _pc_findings.FROM_DATE,
                _pc_findings.THRU_DATE,
                _pc_findings.INPUT_DATE,
                "as_of_date",
                "begin_value_date_a",
                "begin_value_date_b",
                "end_value_date_a",
                "end_value_date_b",
            }:
                cell.number_format = "yyyy-mm-dd"
            elif _is_workbook_numeric_column(column_name) and isinstance(
                cell.value,
                (int, float),
            ):
                cell.number_format = WORKBOOK_NUMBER_FORMAT
        if max_width == 0:
            max_width = min(len(header), 10)
        max_width = max(max_width, _longest_header_word_width(header))
        max_width = max(max_width, _MINIMUM_COLUMN_WIDTHS.get(column_name, 0))
        worksheet.column_dimensions[column_letter].width = min(
            max(max_width + 2, 8),
            36,
        )


def _is_workbook_numeric_column(column_name: str) -> bool:
    """Return whether a workbook column should use the general numeric format."""
    normalized_name = column_name.lower()
    return normalized_name in {
        pc_cols.PORTFOLIO_RETURN,
        pc_cols.SECURITY_RETURN,
        "change",
        "performance_change",
        "estimated_cause_total",
        "unexplained_change",
        "estimated_impact",
        "related_performance_difference",
        _pc_findings.SNAPSHOT_A_VALUE,
        _pc_findings.SNAPSHOT_B_VALUE,
        _pc_findings.DELTA_B_MINUS_A,
        _pc_findings.TRANSACTION_IMPACT_DIAGNOSTIC_ESTIMATE,
        _pc_explain.PORTFOLIO_RETURN_DELTA,
        _pc_explain.SECURITY_RETURN_DELTA,
        _pc_explain.ESTIMATED_RETURN_IMPACT,
        _pc_explain.ESTIMATED_RETURN_IMPACT_TOTAL,
        "reported_return_a",
        "reported_return_b",
        "reported_return_difference",
        "derived_return_a",
        "derived_return_b",
        "derived_return_difference",
        "reconstruction_difference",
        "derived_numerator_a",
        "derived_numerator_b",
        "derived_numerator_difference",
        "derived_denominator_a",
        "derived_denominator_b",
        "derived_denominator_difference",
        "begin_value_a",
        "begin_value_b",
        "begin_value_difference",
        "end_value_a",
        "end_value_b",
        "end_value_difference",
        "net_flow_a",
        "net_flow_b",
        "net_flow_difference",
        "weighted_flow_a",
        "weighted_flow_b",
        "weighted_flow_difference",
        "income_a",
        "income_b",
        "income_difference",
        "row_count",
    }


def _review_workbook_sheet_issues(workbook: Any) -> list[str]:
    """Return review workbook sheet and header validation issues."""
    issues: list[str] = []
    sheet_names = tuple(str(name) for name in workbook.sheetnames)
    if _pc_review_model.PERFORMANCE_DIFFERENCES_SHEET not in sheet_names:
        issues.append(
            "report.xlsx is missing primary sheet "
            f"{_pc_review_model.PERFORMANCE_DIFFERENCES_SHEET!r}"
        )
    required_sheets: list[str] = list(SHARED_REVIEW_SHEETS)
    for sheet_name in (*PRIMARY_DIFFERENCE_SHEETS, *required_sheets):
        if sheet_name not in sheet_names:
            if sheet_name in SHARED_REVIEW_SHEETS:
                issues.append(f"report.xlsx is missing sheet {sheet_name!r}")
            continue
        issues.extend(_review_workbook_header_issues(workbook[sheet_name], sheet_name))
    return issues


def _review_workbook_header_issues(worksheet: Any, sheet_name: str) -> list[str]:
    """Return header validation issues for one review workbook sheet."""
    rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        headers = tuple(str(value) for value in next(rows) if value is not None)
    except StopIteration:
        return [f"report.xlsx sheet {sheet_name!r} has no header row"]

    normalized_headers = tuple(_normalize_header(value) for value in headers)
    missing_headers = [
        header
        for header in REQUIRED_HEADERS[sheet_name]
        if header not in normalized_headers
    ]
    if not missing_headers:
        return []
    return [
        f"report.xlsx sheet {sheet_name!r} is missing headers "
        f"{missing_headers}"
    ]


def _format_value(value: object) -> str:
    """Return a compact display value for workbook sizing."""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10g}"
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return str(value)


def _normalize_header(value: object) -> str:
    """Return a header value with Excel line breaks normalized to spaces."""
    return " ".join(str(value).split())


def _longest_header_word_width(header: str) -> int:
    """Return the minimum width needed to avoid splitting header words."""
    words = header.replace("\n", " ").split()
    if not words:
        return 0
    return max(len(word) for word in words) + 2
