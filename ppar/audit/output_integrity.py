"""Canonical report semantics and generated-bundle integrity validation."""

from __future__ import annotations

# Python imports
from collections.abc import Mapping, Sequence
import datetime as dt
import hashlib
from html.parser import HTMLParser
import io
import json
from pathlib import Path
from typing import Any, TypeGuard

# Third-party imports
import polars as pl

# Project imports
from ppar.audit import rendering as _pc_rendering
from ppar.audit import review_model as _pc_review_model
from ppar.audit import workbook as _pc_workbook

NORMALIZATION_VERSION = 1
VOLATILE_METADATA = (
    "manifest.created_at",
    "xlsx.core_properties.created",
    "xlsx.core_properties.modified",
    "xlsx.zip_entry_timestamps",
)
_BUNDLE_FINGERPRINT = "bundle_fingerprint"
_DISPLAY_HASH = "display_sha256"
_SEMANTIC_HASH = "semantic_sha256"
_REVIEW_KEY_COLUMNS = {"review_key", "reconstruction_review_key"}


def table_manifest_metadata(table: pl.DataFrame) -> dict[str, object]:
    """Return deterministic manifest metadata for one persisted CSV table.

    Args:
        table: Internal table before CSV serialization.

    Returns:
        Row count, ordered columns, and normalized semantic fingerprint.
    """
    column_types = [_dtype_kind(data_type) for data_type in table.dtypes]
    return {
        "rows": table.height,
        "columns": list(table.columns),
        "column_types": column_types,
        _SEMANTIC_HASH: _generic_table_hash(table, column_types),
    }


def review_sheet_manifest_metadata(
    sheets: Sequence[_pc_workbook.ReviewWorkbookSheet],
) -> dict[str, object]:
    """Return canonical display metadata for workbook-style review sheets.

    Args:
        sheets: Internal sheet specifications shared by HTML and XLSX output.

    Returns:
        Metadata keyed by stable review-sheet artifact name.
    """
    metadata: dict[str, object] = {}
    for sheet in sheets:
        columns = _review_sheet_columns(sheet)
        headers = _review_sheet_headers(sheet, columns)
        metadata[sheet.artifact_name] = {
            "sheet_name": sheet.sheet_name,
            "rows": sheet.table.height,
            "internal_columns": list(columns),
            "display_headers": headers,
            _DISPLAY_HASH: _display_table_hash(sheet.table, columns, headers),
        }
    return metadata


def output_integrity_metadata(
    sheets: Sequence[_pc_workbook.ReviewWorkbookSheet],
) -> dict[str, object]:
    """Return the nonvolatile output-integrity manifest section."""
    return {
        "normalization_version": NORMALIZATION_VERSION,
        "volatile_metadata": list(VOLATILE_METADATA),
        "review_sheets": review_sheet_manifest_metadata(sheets),
    }


def with_normalized_bundle_fingerprint(
    manifest: Mapping[str, object],
) -> dict[str, object]:
    """Return a manifest copy containing its normalized bundle fingerprint."""
    normalized = {str(key): value for key, value in manifest.items()}
    integrity = _mapping(normalized.get("output_integrity"))
    integrity[_BUNDLE_FINGERPRINT] = normalized_manifest_fingerprint(normalized)
    normalized["output_integrity"] = integrity
    return normalized


def normalized_manifest_fingerprint(manifest: Mapping[str, object]) -> str:
    """Return a stable bundle fingerprint after removing declared volatility.

    The fingerprint covers normalized manifest semantics, including ordered
    table and review-sheet fingerprints. It deliberately excludes only the
    generation timestamp and its own value.
    """
    payload = {str(key): value for key, value in manifest.items() if key != "created_at"}
    integrity = _mapping(payload.get("output_integrity"))
    integrity.pop(_BUNDLE_FINGERPRINT, None)
    payload["output_integrity"] = integrity
    return _payload_hash(payload)


def report_bundle_output_integrity_issues(
    bundle_path: Path,
    manifest: Mapping[str, object],
    artifacts: Mapping[str, object],
    tables: Mapping[str, object],
) -> list[str]:
    """Return SN-10/SN-11 semantic parity and determinism issues.

    Args:
        bundle_path: Root directory of the generated bundle.
        manifest: Parsed bundle manifest.
        artifacts: Manifest artifact mapping.
        tables: Manifest table metadata mapping.

    Returns:
        Human-readable validation issues. An empty list means output integrity
        is valid.
    """
    integrity = manifest.get("output_integrity")
    if not isinstance(integrity, dict):
        return ["manifest output_integrity is missing or malformed"]

    issues = _output_integrity_shape_issues(integrity)
    expected_fingerprint = integrity.get(_BUNDLE_FINGERPRINT)
    actual_fingerprint = normalized_manifest_fingerprint(manifest)
    if expected_fingerprint != actual_fingerprint:
        issues.append("manifest normalized bundle fingerprint does not match")
    issues.extend(_persisted_csv_semantic_issues(bundle_path, artifacts, tables))
    review_sheets = _mapping(integrity.get("review_sheets"))
    issues.extend(_html_review_parity_issues(bundle_path, artifacts, review_sheets))
    issues.extend(_xlsx_review_parity_issues(bundle_path, artifacts, review_sheets))
    return issues


def _output_integrity_shape_issues(integrity: Mapping[str, object]) -> list[str]:
    """Return malformed output-integrity metadata issues."""
    issues: list[str] = []
    if integrity.get("normalization_version") != NORMALIZATION_VERSION:
        issues.append("manifest output_integrity normalization_version is unsupported")
    if integrity.get("volatile_metadata") != list(VOLATILE_METADATA):
        issues.append("manifest output_integrity volatile_metadata is malformed")
    review_sheets = integrity.get("review_sheets")
    if not isinstance(review_sheets, dict):
        issues.append("manifest output_integrity review_sheets is malformed")
    else:
        for artifact_name, metadata in review_sheets.items():
            issues.extend(_review_sheet_metadata_issues(str(artifact_name), metadata))
    fingerprint = integrity.get(_BUNDLE_FINGERPRINT)
    if not isinstance(fingerprint, str) or len(fingerprint) != 64:
        issues.append("manifest output_integrity bundle_fingerprint is malformed")
    return issues


def _review_sheet_metadata_issues(
    artifact_name: str,
    raw_metadata: object,
) -> list[str]:
    """Return malformed canonical review-sheet metadata issues."""
    metadata = _mapping(raw_metadata)
    row_count = metadata.get("rows")
    valid = (
        isinstance(metadata.get("sheet_name"), str)
        and isinstance(row_count, int)
        and not isinstance(row_count, bool)
        and row_count >= 0
        and _is_string_list(metadata.get("internal_columns"))
        and _is_string_list(metadata.get("display_headers"))
        and isinstance(metadata.get(_DISPLAY_HASH), str)
        and len(str(metadata.get(_DISPLAY_HASH))) == 64
    )
    if valid:
        return []
    return [f"review sheet {artifact_name!r} metadata is malformed"]


def _persisted_csv_semantic_issues(
    bundle_path: Path,
    artifacts: Mapping[str, object],
    tables: Mapping[str, object],
) -> list[str]:
    """Return issues when a CSV no longer matches its internal table fingerprint."""
    issues: list[str] = []
    for table_name, raw_metadata in tables.items():
        metadata = _mapping(raw_metadata)
        expected_hash = metadata.get(_SEMANTIC_HASH)
        expected_columns = metadata.get("columns")
        column_types = metadata.get("column_types")
        if (
            not isinstance(expected_hash, str)
            or not _is_string_list(expected_columns)
            or not _is_string_list(column_types)
        ):
            issues.append(f"manifest table {table_name!r} semantic metadata is malformed")
            continue
        artifact_name = "findings" if table_name == "findings" else table_name
        artifact_file = artifacts.get(artifact_name)
        if not isinstance(artifact_file, str):
            continue
        csv_path = bundle_path / artifact_file
        if not csv_path.is_file():
            continue
        try:
            # Read the serialized representation without type inference. A value
            # such as a string tolerance of ``0`` must not become the float
            # ``0.0`` merely because every populated CSV value looks numeric.
            table = pl.read_csv(csv_path, infer_schema=False)
        except (OSError, pl.exceptions.PolarsError):
            continue
        actual_hash = _generic_table_hash(table, column_types)
        if table.columns != expected_columns:
            issues.append(f"table {table_name!r} ordered columns do not match manifest")
        if actual_hash != expected_hash:
            issues.append(f"table {table_name!r} semantic fingerprint does not match")
    return issues


def _html_review_parity_issues(
    bundle_path: Path,
    artifacts: Mapping[str, object],
    review_sheets: Mapping[str, object],
) -> list[str]:
    """Return issues when HTML review tables differ from canonical semantics."""
    artifact_file = artifacts.get(_pc_review_model.HTML_REPORT_ARTIFACT)
    if not isinstance(artifact_file, str):
        return []
    html_path = bundle_path / artifact_file
    if not html_path.is_file():
        return []
    parser = _ReviewHtmlParser()
    try:
        parser.feed(html_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError):
        return []

    issues: list[str] = []
    for artifact_name, raw_metadata in review_sheets.items():
        metadata = _mapping(raw_metadata)
        sheet_name = metadata.get("sheet_name")
        expected_rows = metadata.get("rows")
        if not isinstance(sheet_name, str) or not isinstance(expected_rows, int):
            issues.append(f"review sheet {artifact_name!r} metadata is malformed")
            continue
        if sheet_name not in parser.section_titles:
            issues.append(f"HTML report is missing review section {sheet_name!r}")
            continue
        if expected_rows == 0:
            if sheet_name in parser.tables:
                issues.append(f"HTML empty review section {sheet_name!r} contains rows")
            continue
        payload = parser.tables.get(sheet_name)
        if payload is None:
            issues.append(f"HTML report is missing review table {sheet_name!r}")
            continue
        if _payload_hash(payload) != metadata.get(_DISPLAY_HASH):
            issues.append(f"HTML review table {sheet_name!r} parity failed")
    return issues


def _xlsx_review_parity_issues(
    bundle_path: Path,
    artifacts: Mapping[str, object],
    review_sheets: Mapping[str, object],
) -> list[str]:
    """Return issues when an optional XLSX differs from canonical semantics."""
    artifact_file = artifacts.get(_pc_review_model.REVIEW_WORKBOOK_ARTIFACT)
    if artifact_file is None:
        return []
    if not isinstance(artifact_file, str):
        return []
    workbook_path = bundle_path / artifact_file
    if not workbook_path.is_file():
        return []
    try:
        # pylint: disable=import-outside-toplevel
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        return []
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:  # pylint: disable=broad-exception-caught
        return []

    issues: list[str] = []
    for raw_metadata in review_sheets.values():
        metadata = _mapping(raw_metadata)
        sheet_name = metadata.get("sheet_name")
        if not isinstance(sheet_name, str):
            continue
        if sheet_name not in workbook.sheetnames:
            issues.append(f"XLSX report is missing review sheet {sheet_name!r}")
            continue
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows())
        if not rows:
            issues.append(f"XLSX review sheet {sheet_name!r} has no header row")
            continue
        headers = [_normalize_excel_header(cell.value) for cell in rows[0]]
        body = [
            [_xlsx_display_value(cell) for cell in row]
            for row in rows[1:]
            if any(cell.value is not None for cell in row)
        ]
        payload = {"columns": headers, "rows": body}
        if _payload_hash(payload) != metadata.get(_DISPLAY_HASH):
            issues.append(f"XLSX review sheet {sheet_name!r} parity failed")
    workbook.close()
    return issues


def _review_sheet_columns(
    sheet: _pc_workbook.ReviewWorkbookSheet,
) -> tuple[str, ...]:
    """Return ordered visible internal columns for one review sheet."""
    requested = sheet.columns or tuple(sheet.table.columns)
    return tuple(
        column
        for column in requested
        if column in sheet.table.columns and column not in _REVIEW_KEY_COLUMNS
    )


def _review_sheet_headers(
    sheet: _pc_workbook.ReviewWorkbookSheet,
    columns: Sequence[str],
) -> list[str]:
    """Return display headers for one review sheet."""
    labels = sheet.labels or {}
    return [labels.get(column, _pc_rendering.display_header(column)) for column in columns]


def _generic_table_hash(
    table: pl.DataFrame,
    column_types: Sequence[str],
) -> str:
    """Return a vectorized semantic hash stable across CSV serialization."""
    normalized_columns: list[pl.Expr] = []
    for column, kind in zip(table.columns, column_types, strict=True):
        value = pl.col(column)
        if kind == "boolean":
            normalized = value.cast(pl.String).str.strip_chars().str.to_lowercase()
        elif kind == "float":
            numeric = value.cast(pl.Float64, strict=False)
            normalized = (
                pl.when(numeric == 0)
                .then(pl.lit("0.0"))
                .otherwise(numeric.cast(pl.String))
            )
        elif kind == "integer":
            normalized = (
                value.cast(pl.Float64, strict=False)
                .cast(pl.Int64, strict=False)
                .cast(pl.String)
            )
        else:
            normalized = value.cast(pl.String)
        normalized_columns.append(normalized.fill_null("").alias(column))

    normalized_table = table.select(normalized_columns)
    serialized = io.BytesIO()
    normalized_table.write_csv(serialized, include_header=False)
    digest = hashlib.sha256()
    digest.update(
        json.dumps(
            {
                "columns": list(table.columns),
                "column_types": list(column_types),
            },
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8")
    )
    digest.update(b"\n")
    digest.update(serialized.getbuffer())
    return digest.hexdigest()


def _display_table_hash(
    table: pl.DataFrame,
    columns: Sequence[str],
    headers: Sequence[str],
) -> str:
    """Return the display hash without materializing a second full table copy.

    Notes:
        The byte stream is identical to hashing the canonical ``columns`` and
        ``rows`` payload as one JSON object. Streaming rows prevents manifest
        creation from retaining every reviewer-visible cell again at large sites.
    """
    digest = hashlib.sha256()
    digest.update(b'{"columns":')
    digest.update(
        json.dumps(
            list(headers), ensure_ascii=False, separators=(",", ":"), sort_keys=True
        ).encode("utf-8")
    )
    digest.update(b',"rows":[')
    first_row = True
    for row in table.iter_rows(named=True):
        if not first_row:
            digest.update(b",")
        first_row = False
        values = [_display_value(row[column]) for column in columns]
        digest.update(
            json.dumps(
                values, ensure_ascii=False, separators=(",", ":"), sort_keys=True
            ).encode("utf-8")
        )
    digest.update(b"]}")
    return digest.hexdigest()


def _dtype_kind(data_type: pl.DataType) -> str:
    """Return a stable coarse type used for CSV round-trip normalization."""
    if data_type.is_float():
        return "float"
    if data_type.is_integer():
        return "integer"
    if data_type == pl.Boolean:
        return "boolean"
    if data_type == pl.Date:
        return "date"
    if data_type == pl.Datetime:
        return "datetime"
    return "string"


def _display_value(value: object) -> str:
    """Return the canonical reviewer-visible scalar representation."""
    return _pc_rendering.format_value(value)


def _xlsx_display_value(cell: Any) -> str:
    """Return one XLSX cell as the reviewer sees its semantic value."""
    value = cell.value
    if value is None:
        normalized = ""
    elif isinstance(value, bool):
        normalized = "yes" if value else "no"
    elif isinstance(value, dt.datetime):
        if value.time() == dt.time():
            normalized = value.date().isoformat()
        else:
            normalized = value.isoformat()
    elif isinstance(value, dt.date):
        normalized = value.isoformat()
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        if cell.number_format == _pc_workbook.WORKBOOK_NUMBER_FORMAT:
            normalized = f"{float(value):.6f}"
        else:
            normalized = str(value)
    else:
        normalized = str(value)
    return normalized


def _normalize_excel_header(value: object) -> str:
    """Return an XLSX header with line breaks normalized to spaces."""
    return " ".join(str(value or "").split())


def _payload_hash(payload: object) -> str:
    """Return a SHA-256 hash for canonical JSON data."""
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _mapping(value: object) -> dict[str, object]:
    """Return a shallow string-keyed mapping or an empty mapping."""
    if not isinstance(value, dict):
        return {}
    return {str(key): item for key, item in value.items()}


def _is_string_list(value: object) -> TypeGuard[list[str]]:
    """Return whether a value is a list of strings."""
    return isinstance(value, list) and all(isinstance(item, str) for item in value)


class _ReviewHtmlParser(HTMLParser):
    """Extract workbook-style section and table semantics from generated HTML."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.section_titles: set[str] = set()
        self.tables: dict[str, dict[str, object]] = {}
        self._capture: str | None = None
        self._text: list[str] = []
        self._caption = ""
        self._headers: list[str] = []
        self._row: list[str] = []
        self._rows: list[list[str]] = []
        self._in_body = False

    def handle_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        del attrs
        if tag in {"h2", "caption", "th", "td"}:
            self._capture = tag
            self._text = []
        elif tag == "tbody":
            self._in_body = True
        elif tag == "tr" and self._in_body:
            self._row = []
        elif tag == "table":
            self._caption = ""
            self._headers = []
            self._rows = []

    def handle_data(self, data: str) -> None:
        if self._capture is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == self._capture:
            value = "".join(self._text)
            if tag == "h2":
                self.section_titles.add(value)
            elif tag == "caption":
                self._caption = value
            elif tag == "th":
                self._headers.append(value)
            elif tag == "td":
                self._row.append(value)
            self._capture = None
            self._text = []
        if tag == "tr" and self._in_body and self._row:
            self._rows.append(self._row)
        elif tag == "tbody":
            self._in_body = False
        elif tag == "table" and self._caption:
            self.tables[self._caption] = {
                "columns": self._headers,
                "rows": self._rows,
            }
